"""Build the TheraVoca full forecast model (xlsx).

Run: python scripts/build_forecast_xlsx.py
Output: docs/partner-materials/TheraVoca-Forecast.xlsx

Five sheets:
  1. Assumptions       -- editable input cells driving everything below
  2. Revenue Model     -- therapist subscriptions, Stripe fees, net MRR
  3. COGS              -- per-vendor monthly infrastructure cost
  4. P&L Summary       -- revenue, COGS, gross profit, margin %
  5. 12-Month Forecast -- month-by-month from launch with cumulative

All numbers in COGS / P&L / Forecast tabs are FORMULAS referencing the
Assumptions sheet, so anyone editing inputs sees the model recompute
automatically. Cells the user is expected to edit are highlighted in
amber; computed cells in green; headers in brand green.

EXCLUDES paid-ads marketing budget per founder request.
"""
from __future__ import annotations

import os
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Brand palette mirrored from the React app.
BRAND_GREEN = "2D4A3E"
BRAND_GREEN_LIGHT = "3A5E50"
BRAND_CORAL = "C87965"
INK = "2B2A29"
SUBTLE = "6D6A65"
CARD_BG = "FDFBF7"
LINE = "E8E5DF"
PALE_GREEN = "F2F7F1"
PALE_AMBER = "FDF7EC"
PALE_CORAL = "FBE9E5"
INPUT_FILL = "FDF1D6"   # input cells (user-editable)
COMPUTED_FILL = "EAF4EA"  # computed cells (read-only-ish)

OUT_PATH = os.path.abspath(
    os.path.join(
        os.path.dirname(__file__),
        "..",
        "docs",
        "partner-materials",
        "TheraVoca-Forecast.xlsx",
    )
)
os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)


# ──────────────────────────────────────────────────────────────────────
# Style helpers
# ──────────────────────────────────────────────────────────────────────
THIN = Side(style="thin", color=LINE)
BORDER_ALL = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
FONT_TITLE = Font(name="Calibri", bold=True, size=18, color=BRAND_GREEN)
FONT_H2 = Font(name="Calibri", bold=True, size=12, color=BRAND_GREEN)
FONT_KICKER = Font(name="Calibri", bold=True, size=9, color=BRAND_CORAL)
FONT_BODY = Font(name="Calibri", size=11, color=INK)
FONT_BODY_BOLD = Font(name="Calibri", bold=True, size=11, color=INK)
FONT_INPUT = Font(name="Calibri", size=11, color=INK, bold=True)
FONT_COMPUTED = Font(name="Calibri", size=11, color=BRAND_GREEN)
FONT_NOTE = Font(name="Calibri", italic=True, size=9, color=SUBTLE)

FILL_HEADER = PatternFill("solid", fgColor=BRAND_GREEN)
FILL_KICKER = PatternFill("solid", fgColor=CARD_BG)
FILL_INPUT = PatternFill("solid", fgColor=INPUT_FILL)
FILL_COMPUTED = PatternFill("solid", fgColor=COMPUTED_FILL)
FILL_AMBER = PatternFill("solid", fgColor=PALE_AMBER)
FILL_PALE_GREEN = PatternFill("solid", fgColor=PALE_GREEN)


def header_row(ws, row, labels, start_col=1):
    for i, label in enumerate(labels):
        cell = ws.cell(row=row, column=start_col + i, value=label)
        cell.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = BORDER_ALL


def section_title(ws, row, col, text):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = FONT_TITLE
    return cell


def kicker(ws, row, col, text):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = FONT_KICKER
    return cell


def h2(ws, row, col, text):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = FONT_H2
    return cell


def body(ws, row, col, text, bold=False):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = FONT_BODY_BOLD if bold else FONT_BODY
    return cell


def input_cell(ws, row, col, value, number_format="#,##0"):
    """Editable input. Amber fill. User changes these; the model recomputes."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = FONT_INPUT
    cell.fill = FILL_INPUT
    cell.border = BORDER_ALL
    cell.number_format = number_format
    cell.alignment = Alignment(horizontal="right", vertical="center")
    return cell


def computed_cell(ws, row, col, formula, number_format="$#,##0"):
    """Read-only formula. Green fill. Recomputes from inputs."""
    cell = ws.cell(row=row, column=col, value=formula)
    cell.font = FONT_COMPUTED
    cell.fill = FILL_COMPUTED
    cell.border = BORDER_ALL
    cell.number_format = number_format
    cell.alignment = Alignment(horizontal="right", vertical="center")
    return cell


def note(ws, row, col, text):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = FONT_NOTE
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    return cell


def widen(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ──────────────────────────────────────────────────────────────────────
# Workbook construction
# ──────────────────────────────────────────────────────────────────────
wb = Workbook()

# ============= SHEET 1: ASSUMPTIONS =============
ws = wb.active
ws.title = "Assumptions"
widen(ws, [38, 14, 14, 14, 58])

kicker(ws, 1, 1, "INPUTS  --  EDIT THE AMBER CELLS")
section_title(ws, 2, 1, "TheraVoca Forecast Assumptions")
note(
    ws, 3, 1,
    "Three scenarios across columns B/C/D. Anything amber is editable; "
    "every downstream sheet recomputes from these cells. "
    "Excludes paid-ads marketing budget per founder request.",
)
ws.merge_cells("A3:E3")
ws.row_dimensions[3].height = 32

# Scenario headers row 5
header_row(ws, 5, ["Assumption", "Today (pre-launch)", "Soft launch", "Growth", "Notes"])
ws.row_dimensions[5].height = 22

# ── Section: Therapist economics ──
r = 7
h2(ws, r, 1, "Therapist economics")
r += 1

# Therapist subscription price
body(ws, r, 1, "Therapist monthly price ($)")
input_cell(ws, r, 2, 45, "$#,##0")
input_cell(ws, r, 3, 45, "$#,##0")
input_cell(ws, r, 4, 45, "$#,##0")
note(ws, r, 5, "Live in the codebase as $45/mo after a 30-day free trial.")
r += 1

# Trial length
body(ws, r, 1, "Free-trial length (days)")
input_cell(ws, r, 2, 30, "0")
input_cell(ws, r, 3, 30, "0")
input_cell(ws, r, 4, 30, "0")
note(ws, r, 5, "Trial is 30 days; therapists in trial don't generate revenue.")
r += 1

# Active therapists total (signed up)
body(ws, r, 1, "Active therapists (signed up)")
input_cell(ws, r, 2, 5, "#,##0")
input_cell(ws, r, 3, 50, "#,##0")
input_cell(ws, r, 4, 350, "#,##0")
note(ws, r, 5, "Total therapists with live profiles. Today = real + a few testers.")
r += 1

# % in trial
body(ws, r, 1, "% in free trial right now")
input_cell(ws, r, 2, 1.0, "0%")
input_cell(ws, r, 3, 0.30, "0%")
input_cell(ws, r, 4, 0.10, "0%")
note(ws, r, 5, "Higher early on (everyone's new). Lower at scale (most converted long ago).")
r += 1

# Trial-to-paid conversion %
body(ws, r, 1, "Trial-to-paid conversion %")
input_cell(ws, r, 2, 0.55, "0%")
input_cell(ws, r, 3, 0.60, "0%")
input_cell(ws, r, 4, 0.65, "0%")
note(ws, r, 5, "What share of trialing therapists convert to a paid sub. Industry-standard early-product range is 40-70%.")
r += 1

# Computed: paying therapists (used downstream)
body(ws, r, 1, "  -> Paying therapists (computed)", bold=True)
computed_cell(ws, r, 2, "=ROUND(B9*(1-B10)+B9*B10*B11,0)", "#,##0")
computed_cell(ws, r, 3, "=ROUND(C9*(1-C10)+C9*C10*C11,0)", "#,##0")
computed_cell(ws, r, 4, "=ROUND(D9*(1-D10)+D9*D10*D11,0)", "#,##0")
note(ws, r, 5, "Therapists outside the trial + the converted share of trialing therapists.")
PAYING_ROW = r
r += 2

# ── Section: Patient demand ──
h2(ws, r, 1, "Patient demand")
r += 1

# Patient intakes per month
body(ws, r, 1, "Patient intakes per month")
input_cell(ws, r, 2, 5, "#,##0")
input_cell(ws, r, 3, 100, "#,##0")
input_cell(ws, r, 4, 1000, "#,##0")
note(ws, r, 5, "Completed intake submissions. Soft-launch + growth are TheraVoca targets, not industry averages.")
INTAKES_ROW = r
r += 1

# Avg matches surfaced per intake (always 3 today)
body(ws, r, 1, "Therapists matched per intake")
input_cell(ws, r, 2, 3, "0")
input_cell(ws, r, 3, 3, "0")
input_cell(ws, r, 4, 3, "0")
note(ws, r, 5, "Hard-coded in product today.")
r += 1

# ── Section: Per-unit costs ──
r += 1
h2(ws, r, 1, "Per-unit infrastructure costs")
r += 1

# Stripe %
body(ws, r, 1, "Stripe % fee (per charge)")
input_cell(ws, r, 2, 0.029, "0.0%")
input_cell(ws, r, 3, 0.029, "0.0%")
input_cell(ws, r, 4, 0.029, "0.0%")
note(ws, r, 5, "US card processing fee. Same across scenarios.")
STRIPE_PCT_ROW = r
r += 1

# Stripe flat
body(ws, r, 1, "Stripe flat fee per charge ($)")
input_cell(ws, r, 2, 0.30, "$0.00")
input_cell(ws, r, 3, 0.30, "$0.00")
input_cell(ws, r, 4, 0.30, "$0.00")
note(ws, r, 5, "Same across scenarios.")
STRIPE_FLAT_ROW = r
r += 1

# Claude research-warmup $ per therapist per month
body(ws, r, 1, "Anthropic Claude / paying therapist / mo")
input_cell(ws, r, 2, 0.10, "$0.00")
input_cell(ws, r, 3, 0.10, "$0.00")
input_cell(ws, r, 4, 0.12, "$0.00")
note(ws, r, 5, "Research enrichment warmup + occasional bio drafts. Cached 30 days.")
CLAUDE_PER_THERAPIST_ROW = r
r += 1

# OpenAI embedding $ per intake
body(ws, r, 1, "OpenAI embeddings / intake")
input_cell(ws, r, 2, 0.00002, "$0.00000")
input_cell(ws, r, 3, 0.00002, "$0.00000")
input_cell(ws, r, 4, 0.00002, "$0.00000")
note(ws, r, 5, "Patient open-text + deep-match vectors. Trivially small at this token rate.")
EMBED_PER_INTAKE_ROW = r
r += 1

# Cloudflare Stream delivery $ per intake (rough proxy for landing-page views)
body(ws, r, 1, "Video CDN / intake (proxy)")
input_cell(ws, r, 2, 0.01, "$0.00")
input_cell(ws, r, 3, 0.01, "$0.00")
input_cell(ws, r, 4, 0.01, "$0.00")
note(ws, r, 5, "Cloudflare Stream delivery cost. Scales with landing-page traffic; intakes are a rough proxy.")
CDN_PER_INTAKE_ROW = r
r += 1

# Telnyx SMS $ per intake (5 SMS per match round)
body(ws, r, 1, "SMS variable / intake")
input_cell(ws, r, 2, 0.025, "$0.000")
input_cell(ws, r, 3, 0.025, "$0.000")
input_cell(ws, r, 4, 0.025, "$0.000")
note(ws, r, 5, "~5 SMS per match round (alerts to 3 therapists + 2 patient confirms) at $0.005/SMS US.")
SMS_PER_INTAKE_ROW = r
r += 1

# Resend $ per intake
body(ws, r, 1, "Email / intake (free tier 0 until 3K/mo)")
input_cell(ws, r, 2, 0.0, "$0.000")
input_cell(ws, r, 3, 0.0, "$0.000")
input_cell(ws, r, 4, 0.0004, "$0.000")
note(ws, r, 5, "Resend is free up to 3K/mo. Growth assumes $20/mo / 50K emails.")
EMAIL_PER_INTAKE_ROW = r
r += 2

# ── Section: Fixed costs ──
h2(ws, r, 1, "Fixed costs (don't scale with intakes)")
r += 1

# Render web services
body(ws, r, 1, "Render hosting (web services)")
input_cell(ws, r, 2, 7, "$#,##0")
input_cell(ws, r, 3, 14, "$#,##0")
input_cell(ws, r, 4, 50, "$#,##0")
note(ws, r, 5, "$7 Starter staging; +$7 prod soft launch; prod upgrades to Standard ($25) at growth.")
RENDER_ROW = r
r += 1

# MongoDB Atlas
body(ws, r, 1, "MongoDB Atlas")
input_cell(ws, r, 2, 0, "$#,##0")
input_cell(ws, r, 3, 0, "$#,##0")
input_cell(ws, r, 4, 57, "$#,##0")
note(ws, r, 5, "M0 free covers today + soft launch. M10 ($57) when storage caps OR when we sign a BAA.")
MONGO_ROW = r
r += 1

# Sentry
body(ws, r, 1, "Sentry error tracking")
input_cell(ws, r, 2, 0, "$#,##0")
input_cell(ws, r, 3, 0, "$#,##0")
input_cell(ws, r, 4, 26, "$#,##0")
note(ws, r, 5, "Developer tier free up to 5K errors/mo. Team $26 at growth.")
SENTRY_ROW = r
r += 1

# Telnyx fixed
body(ws, r, 1, "Telnyx SMS (fixed: number + brand + campaign)")
input_cell(ws, r, 2, 5, "$#,##0")
input_cell(ws, r, 3, 15, "$#,##0")
input_cell(ws, r, 4, 25, "$#,##0")
note(ws, r, 5, "Phone number + A2P brand + 10DLC campaign vetting. Higher tier at scale.")
TELNYX_FIXED_ROW = r
r += 1

# Resend fixed
body(ws, r, 1, "Resend (email base plan)")
input_cell(ws, r, 2, 0, "$#,##0")
input_cell(ws, r, 3, 0, "$#,##0")
input_cell(ws, r, 4, 20, "$#,##0")
note(ws, r, 5, "Free up to 3K/mo. $20 for 50K at growth.")
RESEND_FIXED_ROW = r
r += 1

# Domain
body(ws, r, 1, "Domain (theravoca.com amortized)")
input_cell(ws, r, 2, 1, "$#,##0")
input_cell(ws, r, 3, 1, "$#,##0")
input_cell(ws, r, 4, 1, "$#,##0")
note(ws, r, 5, "~$12/yr.")
DOMAIN_ROW = r
r += 1

# Cloudflare Stream fixed (storage only)
body(ws, r, 1, "Cloudflare Stream storage")
input_cell(ws, r, 2, 1, "$#,##0")
input_cell(ws, r, 3, 1, "$#,##0")
input_cell(ws, r, 4, 1, "$#,##0")
note(ws, r, 5, "5 testimonial videos stored. Delivery is variable; covered above per-intake.")
STREAM_FIXED_ROW = r
r += 2

# Reference markers for later sheets:
ws.cell(row=r, column=1).value = ""
ROWS = {
    "paying": PAYING_ROW,
    "intakes": INTAKES_ROW,
    "stripe_pct": STRIPE_PCT_ROW,
    "stripe_flat": STRIPE_FLAT_ROW,
    "claude_per_therapist": CLAUDE_PER_THERAPIST_ROW,
    "embed_per_intake": EMBED_PER_INTAKE_ROW,
    "cdn_per_intake": CDN_PER_INTAKE_ROW,
    "sms_per_intake": SMS_PER_INTAKE_ROW,
    "email_per_intake": EMAIL_PER_INTAKE_ROW,
    "render": RENDER_ROW,
    "mongo": MONGO_ROW,
    "sentry": SENTRY_ROW,
    "telnyx_fixed": TELNYX_FIXED_ROW,
    "resend_fixed": RESEND_FIXED_ROW,
    "domain": DOMAIN_ROW,
    "stream_fixed": STREAM_FIXED_ROW,
    "price": 7,  # therapist monthly price row
}


# ============= SHEET 2: REVENUE MODEL =============
ws = wb.create_sheet("Revenue Model")
widen(ws, [38, 18, 18, 18, 55])

kicker(ws, 1, 1, "REVENUE")
section_title(ws, 2, 1, "Monthly recurring revenue")
note(ws, 3, 1, "All numbers are formulas referencing the Assumptions sheet. Edit Assumptions to recompute.")
ws.merge_cells("A3:E3")

header_row(ws, 5, ["Line item", "Today", "Soft launch", "Growth", "Notes"])

r = 7
body(ws, r, 1, "Paying therapists")
for col, scen_col in enumerate(["B", "C", "D"], start=2):
    computed_cell(
        ws, r, col, f"=Assumptions!{scen_col}{ROWS['paying']}", "#,##0",
    )
note(ws, r, 5, "From Assumptions sheet: signed-up * (1 - trial%) + trial * conversion%.")
PAYING_REVMODEL_ROW = r
r += 1

body(ws, r, 1, "Therapist monthly price")
for col, scen_col in enumerate(["B", "C", "D"], start=2):
    computed_cell(ws, r, col, f"=Assumptions!{scen_col}{ROWS['price']}", "$#,##0")
note(ws, r, 5, "From Assumptions.")
PRICE_REVMODEL_ROW = r
r += 1

body(ws, r, 1, "Gross MRR (therapist subs)", bold=True)
for col in [2, 3, 4]:
    sc = get_column_letter(col)
    computed_cell(
        ws, r, col,
        f"={sc}{PAYING_REVMODEL_ROW}*{sc}{PRICE_REVMODEL_ROW}",
        "$#,##0",
    )
note(ws, r, 5, "Paying therapists * monthly price.")
GROSS_MRR_ROW = r
r += 1

body(ws, r, 1, "Stripe fees")
for col in [2, 3, 4]:
    sc = get_column_letter(col)
    computed_cell(
        ws, r, col,
        (
            f"={sc}{GROSS_MRR_ROW}*Assumptions!{sc}{ROWS['stripe_pct']}"
            f"+{sc}{PAYING_REVMODEL_ROW}*Assumptions!{sc}{ROWS['stripe_flat']}"
        ),
        "$#,##0",
    )
note(ws, r, 5, "2.9% on gross MRR + $0.30 per charge (one per paying therapist).")
STRIPE_FEES_ROW = r
r += 1

body(ws, r, 1, "Net MRR after Stripe", bold=True)
for col in [2, 3, 4]:
    sc = get_column_letter(col)
    computed_cell(
        ws, r, col,
        f"={sc}{GROSS_MRR_ROW}-{sc}{STRIPE_FEES_ROW}",
        "$#,##0",
    )
note(ws, r, 5, "What lands in the business account each month.")
NET_MRR_ROW = r
r += 1


# ============= SHEET 3: COGS / INFRASTRUCTURE =============
ws = wb.create_sheet("COGS")
widen(ws, [38, 18, 18, 18, 55])

kicker(ws, 1, 1, "COGS")
section_title(ws, 2, 1, "Monthly infrastructure cost")
note(
    ws, 3, 1,
    "Per-vendor monthly bill. Variable costs scale with intake volume "
    "and paying therapists; fixed costs are flat at each scenario tier.",
)
ws.merge_cells("A3:E3")

header_row(ws, 5, ["Line item", "Today", "Soft launch", "Growth", "Notes"])

r = 7
h2(ws, r, 1, "Variable")
r += 1

# Claude API
body(ws, r, 1, "Anthropic Claude API")
for col, sc in enumerate(["B", "C", "D"], start=2):
    computed_cell(
        ws, r, col,
        f"=Assumptions!{sc}{ROWS['paying']}*Assumptions!{sc}{ROWS['claude_per_therapist']}",
        "$#,##0.00",
    )
note(ws, r, 5, "Paying therapists * Claude $ per therapist per month (Assumptions).")
r += 1

# OpenAI embeddings
body(ws, r, 1, "OpenAI embeddings")
for col, sc in enumerate(["B", "C", "D"], start=2):
    computed_cell(
        ws, r, col,
        f"=Assumptions!{sc}{ROWS['intakes']}*Assumptions!{sc}{ROWS['embed_per_intake']}",
        "$#,##0.00",
    )
note(ws, r, 5, "Intakes * per-intake embedding cost.")
r += 1

# Cloudflare Stream delivery
body(ws, r, 1, "Cloudflare Stream delivery")
for col, sc in enumerate(["B", "C", "D"], start=2):
    computed_cell(
        ws, r, col,
        f"=Assumptions!{sc}{ROWS['intakes']}*Assumptions!{sc}{ROWS['cdn_per_intake']}",
        "$#,##0.00",
    )
note(ws, r, 5, "Approximated by intakes (proxy for landing-page traffic).")
r += 1

# Telnyx SMS variable
body(ws, r, 1, "Telnyx SMS (variable)")
for col, sc in enumerate(["B", "C", "D"], start=2):
    computed_cell(
        ws, r, col,
        f"=Assumptions!{sc}{ROWS['intakes']}*Assumptions!{sc}{ROWS['sms_per_intake']}",
        "$#,##0.00",
    )
note(ws, r, 5, "Intakes * ~5 SMS each at US carrier rate.")
r += 1

# Email variable
body(ws, r, 1, "Resend email (variable)")
for col, sc in enumerate(["B", "C", "D"], start=2):
    computed_cell(
        ws, r, col,
        f"=Assumptions!{sc}{ROWS['intakes']}*Assumptions!{sc}{ROWS['email_per_intake']}",
        "$#,##0.00",
    )
note(ws, r, 5, "Free under 3K/mo; growth assumes paid plan.")
r += 1

# Sum of variable
body(ws, r, 1, "Subtotal variable", bold=True)
for col in [2, 3, 4]:
    sc = get_column_letter(col)
    computed_cell(ws, r, col, f"=SUM({sc}8:{sc}{r-1})", "$#,##0.00")
SUBTOTAL_VAR_ROW = r
r += 2

h2(ws, r, 1, "Fixed")
r += 1

# Render
body(ws, r, 1, "Render hosting")
for col, sc in enumerate(["B", "C", "D"], start=2):
    computed_cell(ws, r, col, f"=Assumptions!{sc}{ROWS['render']}", "$#,##0")
r += 1

# Mongo
body(ws, r, 1, "MongoDB Atlas")
for col, sc in enumerate(["B", "C", "D"], start=2):
    computed_cell(ws, r, col, f"=Assumptions!{sc}{ROWS['mongo']}", "$#,##0")
r += 1

# Sentry
body(ws, r, 1, "Sentry")
for col, sc in enumerate(["B", "C", "D"], start=2):
    computed_cell(ws, r, col, f"=Assumptions!{sc}{ROWS['sentry']}", "$#,##0")
r += 1

# Telnyx fixed
body(ws, r, 1, "Telnyx fixed (number + A2P)")
for col, sc in enumerate(["B", "C", "D"], start=2):
    computed_cell(ws, r, col, f"=Assumptions!{sc}{ROWS['telnyx_fixed']}", "$#,##0")
r += 1

# Resend fixed
body(ws, r, 1, "Resend base")
for col, sc in enumerate(["B", "C", "D"], start=2):
    computed_cell(ws, r, col, f"=Assumptions!{sc}{ROWS['resend_fixed']}", "$#,##0")
r += 1

# Domain
body(ws, r, 1, "Domain")
for col, sc in enumerate(["B", "C", "D"], start=2):
    computed_cell(ws, r, col, f"=Assumptions!{sc}{ROWS['domain']}", "$#,##0")
r += 1

# Stream storage
body(ws, r, 1, "Cloudflare Stream storage")
for col, sc in enumerate(["B", "C", "D"], start=2):
    computed_cell(ws, r, col, f"=Assumptions!{sc}{ROWS['stream_fixed']}", "$#,##0")
r += 1

# Subtotal fixed
body(ws, r, 1, "Subtotal fixed", bold=True)
fixed_start = SUBTOTAL_VAR_ROW + 3
for col in [2, 3, 4]:
    sc = get_column_letter(col)
    computed_cell(ws, r, col, f"=SUM({sc}{fixed_start}:{sc}{r-1})", "$#,##0")
SUBTOTAL_FIXED_ROW = r
r += 2

# Total COGS
body(ws, r, 1, "TOTAL monthly infrastructure", bold=True)
for col in [2, 3, 4]:
    sc = get_column_letter(col)
    cell = computed_cell(
        ws, r, col,
        f"={sc}{SUBTOTAL_VAR_ROW}+{sc}{SUBTOTAL_FIXED_ROW}",
        "$#,##0.00",
    )
    cell.font = Font(name="Calibri", bold=True, size=12, color=BRAND_GREEN)
    cell.fill = PatternFill("solid", fgColor=PALE_GREEN)
TOTAL_COGS_ROW = r


# ============= SHEET 4: P&L SUMMARY =============
ws = wb.create_sheet("P&L Summary")
widen(ws, [38, 18, 18, 18, 55])

kicker(ws, 1, 1, "P&L  --  EXCLUDES PAID-ADS MARKETING")
section_title(ws, 2, 1, "Monthly P&L summary")
note(
    ws, 3, 1,
    "Revenue - COGS = Gross profit. Paid-ad spend is intentionally NOT in "
    "this model -- the founder budgets that separately based on growth targets.",
)
ws.merge_cells("A3:E3")

header_row(ws, 5, ["", "Today", "Soft launch", "Growth", "Notes"])

r = 7
body(ws, r, 1, "Net MRR (after Stripe)", bold=True)
for col, sc in enumerate(["B", "C", "D"], start=2):
    computed_cell(ws, r, col, f"='Revenue Model'!{sc}{NET_MRR_ROW}", "$#,##0")
note(ws, r, 5, "From Revenue Model sheet.")
PNL_REV_ROW = r
r += 1

body(ws, r, 1, "Infrastructure COGS")
for col, sc in enumerate(["B", "C", "D"], start=2):
    computed_cell(ws, r, col, f"=COGS!{sc}{TOTAL_COGS_ROW}", "$#,##0")
note(ws, r, 5, "From COGS sheet.")
PNL_COGS_ROW = r
r += 1

body(ws, r, 1, "Gross profit", bold=True)
for col in [2, 3, 4]:
    sc = get_column_letter(col)
    cell = computed_cell(
        ws, r, col,
        f"={sc}{PNL_REV_ROW}-{sc}{PNL_COGS_ROW}",
        "$#,##0",
    )
    cell.font = Font(name="Calibri", bold=True, size=12, color=BRAND_GREEN)
    cell.fill = PatternFill("solid", fgColor=PALE_GREEN)
note(ws, r, 5, "Net MRR - infrastructure COGS.")
PNL_GP_ROW = r
r += 1

body(ws, r, 1, "Gross margin %")
for col in [2, 3, 4]:
    sc = get_column_letter(col)
    cell = computed_cell(
        ws, r, col,
        (
            f"=IF({sc}{PNL_REV_ROW}=0,0,"
            f"{sc}{PNL_GP_ROW}/{sc}{PNL_REV_ROW})"
        ),
        "0%",
    )
note(ws, r, 5, "Gross profit / net MRR.")
PNL_GM_ROW = r
r += 2

# A few derived "unit economics" rows
h2(ws, r, 1, "Unit economics")
r += 1
body(ws, r, 1, "Gross profit per paying therapist")
for col in [2, 3, 4]:
    sc = get_column_letter(col)
    computed_cell(
        ws, r, col,
        (
            f"=IF('Revenue Model'!{sc}{PAYING_REVMODEL_ROW}=0,0,"
            f"{sc}{PNL_GP_ROW}/'Revenue Model'!{sc}{PAYING_REVMODEL_ROW})"
        ),
        "$#,##0.00",
    )
note(ws, r, 5, "Gross profit / paying therapists. Useful for sizing CAC budget.")
r += 1

body(ws, r, 1, "COGS per intake")
for col in [2, 3, 4]:
    sc = get_column_letter(col)
    computed_cell(
        ws, r, col,
        (
            f"=IF(Assumptions!{sc}{ROWS['intakes']}=0,0,"
            f"{sc}{PNL_COGS_ROW}/Assumptions!{sc}{ROWS['intakes']})"
        ),
        "$#,##0.00",
    )
note(ws, r, 5, "Total COGS divided by intakes. A patient-acquisition cost ceiling on the platform side.")
r += 1


# ============= SHEET 5: 12-MONTH FORECAST =============
ws = wb.create_sheet("12-Month Forecast")
widen(ws, [10, 18, 18, 18, 18, 18, 18])

kicker(ws, 1, 1, "12-MONTH PROJECTION")
section_title(ws, 2, 1, "Monthly trajectory from soft-launch to growth")
note(
    ws, 3, 1,
    "Straight-line interpolation from soft-launch month-1 toward growth "
    "month-12 across two drivers (paying therapists + intakes per month). "
    "Cumulative cash columns assume reinvestment-free P&L "
    "(excludes paid-ad spend, owner draws, taxes).",
)
ws.merge_cells("A3:G3")

header_row(ws, 5, [
    "Month",
    "Paying therapists",
    "Intakes / mo",
    "Net MRR",
    "Monthly COGS",
    "Monthly gross profit",
    "Cumulative GP",
])

# Reference soft-launch and growth values from Assumptions
# We compute paying therapists at month N as: linear from col C to col D over 12 months
# Intakes at month N: linear from col C to col D over 12 months
# COGS uses the same logic via formulas on the per-month row.
SL_PAY = f"'Revenue Model'!C{PAYING_REVMODEL_ROW}"
GR_PAY = f"'Revenue Model'!D{PAYING_REVMODEL_ROW}"
SL_INTAKES = f"Assumptions!C{ROWS['intakes']}"
GR_INTAKES = f"Assumptions!D{ROWS['intakes']}"

# Per-paying-therapist Claude + per-intake variable costs (use growth-tier inputs;
# they're constants and the differences between soft/growth are tiny)
CLAUDE_PER = f"Assumptions!D{ROWS['claude_per_therapist']}"
EMBED_PER = f"Assumptions!D{ROWS['embed_per_intake']}"
CDN_PER = f"Assumptions!D{ROWS['cdn_per_intake']}"
SMS_PER = f"Assumptions!D{ROWS['sms_per_intake']}"
EMAIL_PER = f"Assumptions!D{ROWS['email_per_intake']}"
STRIPE_PCT = f"Assumptions!D{ROWS['stripe_pct']}"
STRIPE_FLAT = f"Assumptions!D{ROWS['stripe_flat']}"
PRICE = f"Assumptions!D{ROWS['price']}"

# Fixed costs slowly grow from soft-launch tier to growth tier across the year.
# We interpolate the SUM of fixed costs the same way.
SL_FIXED_SUM = (
    f"(Assumptions!C{ROWS['render']}+Assumptions!C{ROWS['mongo']}"
    f"+Assumptions!C{ROWS['sentry']}+Assumptions!C{ROWS['telnyx_fixed']}"
    f"+Assumptions!C{ROWS['resend_fixed']}+Assumptions!C{ROWS['domain']}"
    f"+Assumptions!C{ROWS['stream_fixed']})"
)
GR_FIXED_SUM = (
    f"(Assumptions!D{ROWS['render']}+Assumptions!D{ROWS['mongo']}"
    f"+Assumptions!D{ROWS['sentry']}+Assumptions!D{ROWS['telnyx_fixed']}"
    f"+Assumptions!D{ROWS['resend_fixed']}+Assumptions!D{ROWS['domain']}"
    f"+Assumptions!D{ROWS['stream_fixed']})"
)

for m in range(1, 13):
    r = 6 + m  # row 7..18
    # progress fraction 0 (month 1) to 1 (month 12)
    progress = f"(({m}-1)/11)"
    # Paying therapists at this month
    body(ws, r, 1, f"M{m}")
    computed_cell(
        ws, r, 2,
        f"=ROUND({SL_PAY}+({GR_PAY}-{SL_PAY})*{progress},0)",
        "#,##0",
    )
    # Intakes
    computed_cell(
        ws, r, 3,
        f"=ROUND({SL_INTAKES}+({GR_INTAKES}-{SL_INTAKES})*{progress},0)",
        "#,##0",
    )
    # Net MRR: paying * price - (paying * price * stripe_pct + paying * stripe_flat)
    pay_ref = f"B{r}"
    intakes_ref = f"C{r}"
    computed_cell(
        ws, r, 4,
        (
            f"={pay_ref}*{PRICE}"
            f"-({pay_ref}*{PRICE}*{STRIPE_PCT}"
            f"+{pay_ref}*{STRIPE_FLAT})"
        ),
        "$#,##0",
    )
    # Monthly COGS: variable + interpolated fixed
    computed_cell(
        ws, r, 5,
        (
            f"={pay_ref}*{CLAUDE_PER}"
            f"+{intakes_ref}*({EMBED_PER}+{CDN_PER}+{SMS_PER}+{EMAIL_PER})"
            f"+{SL_FIXED_SUM}+({GR_FIXED_SUM}-{SL_FIXED_SUM})*{progress}"
        ),
        "$#,##0",
    )
    # Gross profit
    computed_cell(ws, r, 6, f"=D{r}-E{r}", "$#,##0")
    # Cumulative GP
    if m == 1:
        computed_cell(ws, r, 7, f"=F{r}", "$#,##0")
    else:
        computed_cell(ws, r, 7, f"=G{r-1}+F{r}", "$#,##0")

# Year-1 totals row
total_r = 19
body(ws, total_r, 1, "Year 1", bold=True)
for col in [4, 5, 6]:
    sc = get_column_letter(col)
    cell = computed_cell(ws, total_r, col, f"=SUM({sc}7:{sc}18)", "$#,##0")
    cell.font = Font(name="Calibri", bold=True, size=12, color=BRAND_GREEN)
    cell.fill = PatternFill("solid", fgColor=PALE_GREEN)
# Cumulative GP equals year-1 sum
cell = computed_cell(ws, total_r, 7, f"=G18", "$#,##0")
cell.font = Font(name="Calibri", bold=True, size=12, color=BRAND_GREEN)
cell.fill = PatternFill("solid", fgColor=PALE_GREEN)


# ============= COVER SHEET WITH NAVIGATION =============
# Put a friendly "Read me" sheet at the front
wb.create_sheet("Read me", 0)
ws = wb["Read me"]
widen(ws, [80])

kicker(ws, 1, 1, "TheraVoca FORECAST MODEL")
section_title(ws, 2, 1, "How to use this workbook")

ws.cell(row=4, column=1, value=(
    "This file models TheraVoca's monthly revenue, COGS, and gross "
    "margin across three scenarios:"
)).font = FONT_BODY

ws.cell(row=5, column=1, value=(
    "  - Today (pre-launch testing)"
)).font = FONT_BODY
ws.cell(row=6, column=1, value=(
    "  - Soft launch (~100 patient intakes / month)"
)).font = FONT_BODY
ws.cell(row=7, column=1, value=(
    "  - Growth (~1,000 patient intakes / month)"
)).font = FONT_BODY

ws.cell(row=9, column=1, value=(
    "Edit anything in an AMBER cell on the Assumptions tab; every "
    "downstream sheet recomputes via formulas."
)).font = FONT_BODY

ws.cell(row=11, column=1, value="Sheets in this workbook:").font = FONT_H2

ws.cell(row=13, column=1, value=(
    "Assumptions      -- inputs (price, conversion %, intakes, per-unit costs, fixed costs)"
)).font = FONT_BODY
ws.cell(row=14, column=1, value=(
    "Revenue Model    -- paying therapist count, gross MRR, Stripe fees, net MRR"
)).font = FONT_BODY
ws.cell(row=15, column=1, value=(
    "COGS             -- per-vendor monthly cost broken into variable + fixed"
)).font = FONT_BODY
ws.cell(row=16, column=1, value=(
    "P&L Summary      -- revenue, COGS, gross profit, gross margin %, unit economics"
)).font = FONT_BODY
ws.cell(row=17, column=1, value=(
    "12-Month Forecast -- month-by-month interpolation from soft-launch to growth"
)).font = FONT_BODY

ws.cell(row=19, column=1, value="What's NOT in this model:").font = FONT_H2
ws.cell(row=20, column=1, value=(
    "  - Paid-ads marketing budget (founder budgets separately)"
)).font = FONT_BODY
ws.cell(row=21, column=1, value=(
    "  - Owner draws, salaries, contractor costs"
)).font = FONT_BODY
ws.cell(row=22, column=1, value=(
    "  - Taxes"
)).font = FONT_BODY
ws.cell(row=23, column=1, value=(
    "  - One-time costs (legal, brand registration, etc.)"
)).font = FONT_BODY

ws.cell(row=25, column=1, value=(
    f"Generated {date.today().isoformat()} by Joshua Rosenthal, TheraVoca."
)).font = FONT_NOTE

# Save
wb.save(OUT_PATH)
print(f"Wrote {OUT_PATH}")
