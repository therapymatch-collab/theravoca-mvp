"""Magic-link auth + patient/therapist portal routes."""
from __future__ import annotations

import asyncio
import re
import secrets
import uuid
from datetime import datetime, timedelta, timezone
from typing import Any, Optional

import bcrypt
import json
from fastapi import APIRouter, Depends, HTTPException, Request
from fastapi.responses import Response

from deps import (
    db, logger, MAGIC_CODE_MAX_PER_HOUR, MAGIC_CODE_TTL_MINUTES,
    LOGIN_MAX_FAILURES, LOGIN_LOCKOUT_MINUTES,
    _create_session_token, require_session, _client_ip,
)
from email_service import send_magic_code
import audit
from helpers import _now_iso, _safe_summary_for_therapist, _spawn_bg
from models import MagicCodeRequest, MagicCodeVerify
from profile_completeness import evaluate as _evaluate_profile_inline

router = APIRouter()


# ─── Password helpers ─────────────────────────────────────────────────────────
def _hash_password(plain: str) -> str:
    return bcrypt.hashpw(plain.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")


def _verify_password(plain: str, hashed: str) -> bool:
    if not plain or not hashed:
        return False
    try:
        return bcrypt.checkpw(plain.encode("utf-8"), hashed.encode("utf-8"))
    except (ValueError, TypeError):
        return False


# Persistent (DB-backed) lockout for password login attempts. Keyed by
# `{ip}:{email}` so an attacker can't lock out a real user from another
# IP. Uses the existing `LOGIN_MAX_FAILURES` / `LOGIN_LOCKOUT_MINUTES`
# env constants for parity with the admin login flow.
async def _password_lockout_remaining(key: str) -> Optional[int]:
    rec = await db.password_login_attempts.find_one({"_id": key}, {"_id": 0})
    if not rec or not rec.get("locked_until"):
        return None
    locked_until = rec["locked_until"]
    if isinstance(locked_until, str):
        try:
            locked_until = datetime.fromisoformat(locked_until.replace("Z", "+00:00"))
        except ValueError:
            return None
    now = datetime.now(timezone.utc)
    if locked_until > now:
        return int((locked_until - now).total_seconds())
    await db.password_login_attempts.delete_one({"_id": key})
    return None


async def _password_record_failure(key: str) -> None:
    rec = await db.password_login_attempts.find_one({"_id": key}, {"_id": 0})
    failures = (rec or {}).get("failures", 0) + 1
    update: dict[str, Any] = {"failures": failures, "updated_at": _now_iso()}
    if failures >= LOGIN_MAX_FAILURES:
        update["locked_until"] = (
            datetime.now(timezone.utc) + timedelta(minutes=LOGIN_LOCKOUT_MINUTES)
        ).isoformat()
        logger.warning("Password login locked for key=%s after %d failures", key, failures)
    await db.password_login_attempts.update_one(
        {"_id": key}, {"$set": update}, upsert=True
    )


async def _password_reset_failures(key: str) -> None:
    await db.password_login_attempts.delete_one({"_id": key})


async def _find_user_doc(email: str, role: str) -> Optional[dict[str, Any]]:
    """Locate the auth-bearing record for an email + role.

    Therapists: stored on the `therapists` collection (one doc per therapist).
    Patients:   stored on the `patient_accounts` collection (created lazily
                when the patient sets a password — patients have no global
                profile record otherwise).
    """
    email_norm = email.lower()
    if role == "therapist":
        return await db.therapists.find_one(
            {"email": {"$regex": f"^{re.escape(email_norm)}$", "$options": "i"},
             "is_active": {"$ne": False}},
            {"_id": 0},
        )
    if role == "patient":
        return await db.patient_accounts.find_one({"email": email_norm}, {"_id": 0})
    return None


async def _set_password_on_doc(email: str, role: str, password_hash: str) -> bool:
    """Store the password hash on the appropriate collection. Returns True if
    a record was actually updated/created."""
    email_norm = email.lower()
    now = _now_iso()
    if role == "therapist":
        res = await db.therapists.update_one(
            {"email": {"$regex": f"^{re.escape(email_norm)}$", "$options": "i"}},
            {"$set": {"password_hash": password_hash, "password_set_at": now,
                      "updated_at": now}},
        )
        return res.matched_count > 0
    if role == "patient":
        # Lazy-create a `patient_accounts` row; patients don't have a global
        # profile doc otherwise. We seed `created_at` only on insert.
        await db.patient_accounts.update_one(
            {"email": email_norm},
            {"$set": {"password_hash": password_hash, "password_set_at": now,
                      "updated_at": now},
             "$setOnInsert": {"email": email_norm, "created_at": now}},
            upsert=True,
        )
        return True
    return False



@router.post("/auth/request-code")
async def auth_request_code(payload: MagicCodeRequest, request: Request):
    if payload.role not in ("patient", "therapist"):
        raise HTTPException(400, "Invalid role")
    email = payload.email.lower()

    # Master testing-mode bypass: skip the 5/hour cap so test runs can
    # request as many codes as they need. Honeypot/auth gates remain.
    import testing_mode
    _testing_active = await testing_mode.is_active()
    if not _testing_active:
        one_hour_ago = (datetime.now(timezone.utc) - timedelta(hours=1)).isoformat()
        recent = await db.magic_codes.count_documents({
            "email": email, "role": payload.role, "created_at": {"$gte": one_hour_ago}
        })
        if recent >= MAGIC_CODE_MAX_PER_HOUR:
            raise HTTPException(429, "Too many code requests. Try again in an hour.")
        # SECURITY (2026-05-16 audit, MEDIUM #7): per-IP rate limit.
        # The per-(email, hour) cap above doesn't stop an attacker
        # rotating EMAILS from a single IP -- they could trigger
        # thousands of magic-code emails to arbitrary recipients,
        # damaging Resend deliverability reputation and spamming real
        # inboxes (patient role skips the existence check entirely,
        # so any address goes). Cap per-IP magic-code requests at 20
        # per hour across all roles + emails.
        ip = _client_ip(request)
        if ip:
            ip_recent = await db.magic_codes.count_documents({
                "ip": ip,
                "created_at": {"$gte": one_hour_ago},
            })
            if ip_recent >= 20:
                raise HTTPException(429, "Too many code requests from your network. Try again in an hour.")

    if payload.role == "therapist":
        therapist = await db.therapists.find_one(
            {"email": {"$regex": f"^{re.escape(email)}$", "$options": "i"},
             "is_active": {"$ne": False},
             "pending_approval": {"$ne": True}},
            {"_id": 0, "id": 1},
        )
        if not therapist:
            logger.info("Magic-code requested for unknown/pending therapist")
            return {"ok": True}

    code = f"{secrets.randbelow(900000) + 100000:06d}"
    # Persist the requester IP so the per-IP rate limiter above can
    # count prior requests from the same source. Best-effort -- if we
    # can't resolve an IP (proxy stripped headers) we just store None
    # and the per-IP cap is a no-op for that request.
    requester_ip = _client_ip(request)
    await db.magic_codes.insert_one({
        "id": str(uuid.uuid4()),
        "email": email,
        "role": payload.role,
        "code": code,
        "expires_at": (datetime.now(timezone.utc) + timedelta(minutes=MAGIC_CODE_TTL_MINUTES)).isoformat(),
        "used": False,
        "created_at": _now_iso(),
        "ip": requester_ip,
    })
    _spawn_bg(
        send_magic_code(email, code, payload.role),
        name=f"magic_code_{email[:8]}",
    )
    return {"ok": True}


@router.post("/auth/verify-code")
async def auth_verify_code(payload: MagicCodeVerify, request: Request):
    email = payload.email.lower()
    # Per-(ip, email) failed-attempt lockout. Keyed with a "code:" prefix
    # so it never collides with password login failures (which use the
    # bare ip:email key). Without this, an attacker who can reach our
    # API could brute-force a 6-digit code (1M possibilities) inside the
    # 30-min TTL window.
    ip = _client_ip(request)
    attempt_key = f"code:{ip}:{email}"
    # Master testing-mode bypass: skip the per-(ip, email) lockout so
    # test runs that intentionally probe wrong codes don't get locked.
    import testing_mode
    _testing_active = await testing_mode.is_active()
    if not _testing_active:
        locked = await _password_lockout_remaining(attempt_key)
        if locked is not None:
            raise HTTPException(429, f"Too many failed attempts. Try again in {locked // 60 + 1} min.")

    now_iso = _now_iso()
    rec = await db.magic_codes.find_one({
        "email": email,
        "role": payload.role,
        "code": payload.code,
        "used": False,
        "expires_at": {"$gte": now_iso},
    }, {"_id": 0})
    if not rec:
        # Only record the failure when lockout is enforced -- don't
        # accumulate counts during testing mode.
        if not _testing_active:
            await _password_record_failure(attempt_key)
        raise HTTPException(401, "Invalid or expired code")
    await _password_reset_failures(attempt_key)
    await db.magic_codes.update_one(
        {"id": rec["id"]}, {"$set": {"used": True, "used_at": now_iso}}
    )
    # If the user is a therapist with 2FA enrolled, don't issue the
    # real session token yet -- return a short-lived challenge token
    # the frontend trades for a session at /auth/verify-2fa.
    user = await _find_user_doc(email, payload.role)
    if payload.role == "therapist" and user and user.get("totp_enabled_at"):
        from totp_service import create_challenge_token
        return {
            "requires_2fa": True,
            "challenge_token": create_challenge_token(email, payload.role),
            "role": payload.role,
            "email": email,
        }
    token = _create_session_token(email, payload.role)
    # New-IP login alert (fire-and-forget). Records the event regardless
    # of whether an alert fires; alerts only on second+ login from a
    # never-before-seen IP for this account.
    from login_alerts import check_and_record_login
    await check_and_record_login(
        email=email, role=payload.role,
        ip=_client_ip(request),
        user_agent=request.headers.get("user-agent", ""),
    )
    # Tell the client whether this account already has a password — used by
    # the portal to nudge first-time users to set one for password login.
    has_password = bool(user and user.get("password_hash"))
    return {
        "token": token,
        "role": payload.role,
        "email": email,
        "has_password": has_password,
    }


# ─── Password auth ────────────────────────────────────────────────────────────
@router.get("/auth/password-status")
async def auth_password_status(email: str, role: str):
    """Public — tells the SignIn page whether to show a password input or
    fall back to magic-code only. Always returns 200 to avoid leaking
    account existence; `has_password=False` for unknown emails."""
    if role not in ("patient", "therapist"):
        raise HTTPException(400, "Invalid role")
    user = await _find_user_doc(email, role)
    return {"has_password": bool(user and user.get("password_hash"))}


@router.post("/auth/login-password")
async def auth_login_password(payload: dict[str, Any], request: Request):
    """Email + password login. Returns the same shape as `/auth/verify-code`
    so the frontend can treat both paths interchangeably."""
    email = (payload.get("email") or "").strip().lower()
    password = payload.get("password") or ""
    role = payload.get("role")
    if role not in ("patient", "therapist"):
        raise HTTPException(400, "Invalid role")
    if "@" not in email or not password:
        raise HTTPException(400, "Email and password required")

    ip = _client_ip(request)
    key = f"{ip}:{role}:{email}"
    locked = await _password_lockout_remaining(key)
    if locked is not None:
        raise HTTPException(429, f"Too many failed attempts. Try again in {locked // 60 + 1} min.")

    user = await _find_user_doc(email, role)
    if not user or not user.get("password_hash"):
        await _password_record_failure(key)
        raise HTTPException(401, "Email or password is incorrect")
    if role == "therapist" and user.get("pending_approval"):
        raise HTTPException(403, "Your therapist profile is still under review.")
    if not _verify_password(password, user["password_hash"]):
        await _password_record_failure(key)
        raise HTTPException(401, "Email or password is incorrect")

    await _password_reset_failures(key)
    # 2FA gate (therapists only, opt-in).
    if role == "therapist" and user.get("totp_enabled_at"):
        from totp_service import create_challenge_token
        return {
            "requires_2fa": True,
            "challenge_token": create_challenge_token(email, role),
            "role": role,
            "email": email,
        }
    token = _create_session_token(email, role)
    from login_alerts import check_and_record_login
    await check_and_record_login(
        email=email, role=role,
        ip=ip,
        user_agent=request.headers.get("user-agent", ""),
    )
    return {"token": token, "role": role, "email": email, "has_password": True}


@router.post("/auth/set-password")
async def auth_set_password(
    payload: dict[str, Any],
    session: dict[str, Any] = Depends(require_session(("patient", "therapist"))),
):
    """Set or change the current user's password. Requires an active
    magic-link session (or an existing password session) so we know the
    email is already verified."""
    password = payload.get("password") or ""
    if len(password) < 8:
        raise HTTPException(400, "Password must be at least 8 characters")
    if len(password) > 128:
        raise HTTPException(400, "Password is too long (max 128 chars)")

    ok = await _set_password_on_doc(session["email"], session["role"], _hash_password(password))
    if not ok:
        raise HTTPException(404, "Account not found")
    # Invalidate any prior failure counters so the user can immediately log
    # in with the new password.
    await db.password_login_attempts.delete_many(
        {"_id": {"$regex": f":{re.escape(session['role'])}:{re.escape(session['email'])}$"}}
    )
    return {"ok": True}


@router.get("/portal/me")
async def portal_me(
    session: dict[str, Any] = Depends(require_session(("patient", "therapist"))),
):
    return {"email": session["email"], "role": session["role"]}


@router.get("/portal/login-history")
async def portal_login_history(
    session: dict[str, Any] = Depends(require_session(("patient", "therapist"))),
):
    """Return the signed-in user's last 50 login events.

    Backs the patient + therapist 'My sign-in history' page. Returns
    events with `is_new_device` flagged client-side-style on the server
    (first occurrence of a given ip_hash for this account = new device).
    Raw ip_hash is NOT returned -- not useful to the user, just noise.

    Records live in db.login_events with a 90-day TTL. After 90 days
    a previously-seen IP looks 'new' again, which is intentional.
    """
    from login_alerts import get_login_history
    events = await get_login_history(session["email"], limit=50)
    # Annotate each row with is_new_device. We need ip_hash for the
    # detection but strip it before responding.
    seen_hashes: set[str] = set()
    full_cur = db.login_events.find(
        {"email": session["email"]},
        {"_id": 0, "ip_hash": 1, "ts": 1, "user_agent": 1, "role": 1},
    ).sort("ts", 1)  # ascending so we mark FIRST occurrence as new
    annotated: list[dict] = []
    async for ev in full_cur:
        h = ev.get("ip_hash") or ""
        is_new = bool(h) and h not in seen_hashes
        if h:
            seen_hashes.add(h)
        annotated.append({
            "ts": ev.get("ts"),
            "role": ev.get("role"),
            "user_agent": ev.get("user_agent") or "",
            "is_new_device": is_new,
        })
    # Sort newest-first, cap at 50.
    annotated.sort(key=lambda e: e.get("ts") or "", reverse=True)
    annotated = annotated[:50]
    return {"events": annotated, "count": len(annotated)}


@router.get("/portal/therapist/profile")
async def portal_therapist_profile(
    session: dict[str, Any] = Depends(require_session(("therapist",))),
):
    """Return the therapist's own profile for the self-edit page.

    Strips fields the therapist shouldn't see/own (e.g., internal scoring
    caches) but keeps everything they're allowed to edit. Adds a
    `completeness` block driven by profile_completeness.evaluate()
    so the edit form can render required-field indicators (red asterisks
    + go-live banner) without re-implementing the rules client-side."""
    t = await db.therapists.find_one(
        {"email": {"$regex": f"^{re.escape(session['email'])}$", "$options": "i"}},
        {"_id": 0, "password_hash": 0, "password_set_at": 0,
         "t5_embedding": 0, "t6b_embedding": 0},
    )
    if not t:
        raise HTTPException(404, "Therapist profile not found")
    # Inline the completeness check so the edit form can render
    # required-field indicators. Same evaluator the admin dashboard
    # + claim-profile email use, so the truth source is shared.
    t["completeness"] = _evaluate_profile_inline(t)
    return t


# Fields a therapist can change directly without admin re-approval.
_SELF_EDITABLE_FIELDS = {
    "bio",
    "office_phone", "phone", "phone_alert",
    "website",
    "office_addresses", "office_locations",
    "client_types", "age_groups",
    "modalities", "modality_offering",
    "offers_in_person", "telehealth",
    "insurance_accepted",
    "cash_rate", "sliding_scale", "free_consult",
    "languages_spoken",
    "availability", "availability_notes", "availability_windows",
    "profile_picture",
    "notify_by_email", "notify_by_sms",
    # Deep-match style answers (v5). Therapists fill or update these
    # without admin re-approval — they're style answers, not licensure
    # claims. T5/T6b changes also trigger an embedding refresh.
    "t1_stuck_ranked", "t3_breakthrough",
    "t4_hard_truth", "t5_lived_experience",
    "t6_session_expectations", "t6_early_sessions_description",
}

# Fields that force a re-approval flag when edited (e.g., license or
# specialties changes need a human-in-the-loop check).
_REAPPROVAL_FIELDS = {
    "primary_specialties",
    "secondary_specialties",
    "general_treats",
    "license_number",
    "license_expires_at",
    "licensed_states",
    "years_experience",
    "credential_type",
    "gender",
    "name",
}


@router.put("/portal/therapist/profile")
async def portal_therapist_update_profile(
    payload: dict[str, Any],
    session: dict[str, Any] = Depends(require_session(("therapist",))),
):
    """Therapist self-edit. Silently drops unknown/forbidden fields so
    clients can send the whole profile back without cherry-picking. Setting
    any field in `_REAPPROVAL_FIELDS` flips `pending_reapproval=True` so
    an admin sees it in the pending queue before the change goes live to
    patients."""
    allowed = _SELF_EDITABLE_FIELDS | _REAPPROVAL_FIELDS
    clean = {k: v for k, v in (payload or {}).items() if k in allowed}
    if not clean:
        raise HTTPException(400, "No editable fields provided")

    # Clamp rate to sanity bounds (mirrors models.TherapistSignup)
    if "cash_rate" in clean:
        try:
            clean["cash_rate"] = max(0, min(1000, int(clean["cash_rate"])))
        except (TypeError, ValueError):
            clean.pop("cash_rate")
    # Enforce 3-age-group cap on self-edit too
    if "age_groups" in clean and isinstance(clean["age_groups"], list):
        clean["age_groups"] = clean["age_groups"][:3]

    current = await db.therapists.find_one(
        {"email": {"$regex": f"^{re.escape(session['email'])}$", "$options": "i"}},
        {"_id": 0},
    )
    if not current:
        raise HTTPException(404, "Therapist profile not found")

    reapproval_changed = [
        k for k in _REAPPROVAL_FIELDS
        if k in clean and current.get(k) != clean.get(k)
    ]
    update_doc = {**clean, "updated_at": _now_iso()}
    # Reset the stale-profile nag flag whenever they touch the profile —
    # they're active again and shouldn't get the nag email next week.
    unset_doc = {"stale_profile_nag_sent_at": ""}
    if reapproval_changed:
        update_doc["pending_reapproval"] = True
        update_doc["pending_reapproval_fields"] = reapproval_changed
        update_doc["pending_reapproval_at"] = _now_iso()

    await db.therapists.update_one(
        {"id": current["id"]},
        {"$set": update_doc, "$unset": unset_doc},
    )
    # Re-embed T5 in the background if it changed. We schedule
    # rather than await so the portal-save UX stays snappy.
    t5_changed = "t5_lived_experience" in clean and current.get("t5_lived_experience") != clean.get("t5_lived_experience")
    if t5_changed:
        from routes.therapists import _embed_therapist_signals
        new_t5 = clean.get("t5_lived_experience") if "t5_lived_experience" in clean else current.get("t5_lived_experience", "")
        _spawn_bg(
            _embed_therapist_signals(current["id"], new_t5 or "", ""),
            name=f"embed_portal_{current['id'][:8]}",
        )
    updated = await db.therapists.find_one({"id": current["id"]}, {"_id": 0})
    return {
        "ok": True,
        "profile": updated,
        "requires_reapproval": bool(reapproval_changed),
        "reapproval_fields": reapproval_changed,
    }


@router.get("/portal/patient/requests")
async def portal_patient_requests(
    request: Request,
    session: dict[str, Any] = Depends(require_session(("patient",))),
):
    audit.emit(
        actor_type="patient",
        actor_id=audit._hash_patient_email(session["email"]),
        action="list_own_requests", resource="request",
        ip=request.headers.get("x-forwarded-for", ""),
        user_agent=request.headers.get("user-agent", ""),
    )
    docs = await db.requests.find(
        {"email": {"$regex": f"^{re.escape(session['email'])}$", "$options": "i"}},
        {"_id": 0, "verification_token": 0},
    ).sort("created_at", -1).to_list(100)
    out = []
    for d in docs:
        app_count = await db.applications.count_documents({"request_id": d["id"]})
        d["application_count"] = app_count
        d["notified_count"] = len(d.get("notified_therapist_ids") or [])
        out.append(d)
    # Surface password-account flag so the portal can prompt the patient to
    # set a password if they're still on magic-code-only.
    user = await _find_user_doc(session["email"], "patient")
    # paused_at lives on patient_accounts; surface it so the portal
    # knows whether to render the Pause or Resume CTA. Fetched
    # separately from `user` because _find_user_doc returns the auth
    # row only, not the patient_accounts profile row.
    pa = await db.patient_accounts.find_one(
        {"email": session["email"].lower()},
        {"_id": 0, "paused_at": 1, "deleted_at": 1},
    )
    return {
        "requests": out,
        "has_password": bool(user and user.get("password_hash")),
        "email": session["email"],
        "paused_at": (pa or {}).get("paused_at"),
        "deleted_at": (pa or {}).get("deleted_at"),
    }


@router.get("/portal/therapist/analytics")
async def portal_therapist_analytics(
    request: Request,
    session: dict[str, Any] = Depends(require_session(("therapist",))),
):
    """Lightweight analytics for the therapist portal: how many referrals
    we've sent them, how many they applied to / declined, conversion rate,
    avg match score, top specialty fits, and review summary."""
    from collections import Counter

    audit.emit(
        actor_type="therapist", actor_id=audit._hash_patient_email(session.get("email", "")),
        action="view_analytics", resource="therapist",
        ip=request.headers.get("x-forwarded-for", ""),
        user_agent=request.headers.get("user-agent", ""),
    )
    therapist = await db.therapists.find_one(
        {"email": {"$regex": f"^{re.escape(session['email'])}$", "$options": "i"}},
        {"_id": 0},
    )
    if not therapist:
        raise HTTPException(404, "Therapist profile not found")
    tid = therapist["id"]

    reqs = await db.requests.find(
        {"notified_therapist_ids": tid},
        {"_id": 0, "id": 1, "notified_scores": 1, "presenting_issues": 1,
         "created_at": 1, "results_sent_at": 1},
    ).to_list(500)
    invited = len(reqs)
    score_sum = 0.0
    score_count = 0
    issues_seen: Counter = Counter()
    for r in reqs:
        s = (r.get("notified_scores") or {}).get(tid)
        if s is not None:
            score_sum += float(s)
            score_count += 1
        for i in (r.get("presenting_issues") or []):
            issues_seen[i] += 1
    # Round to whole numbers — the portal UI shows these as `${n}%`
    # chips and decimal precision (85.3%, 76.5%) reads as false-precision
    # to therapists. Whole numbers convey the same information with less
    # cognitive load.
    avg_score = int(round(score_sum / score_count)) if score_count else 0
    applied = await db.applications.count_documents({"therapist_id": tid})
    declined = await db.declines.count_documents({"therapist_id": tid})
    apply_rate = int(round(applied / invited * 100)) if invited else 0
    decline_rate = int(round(declined / invited * 100)) if invited else 0

    # Refer-a-colleague chain
    referrals_made = await db.therapists.count_documents(
        {"referred_by_code": therapist.get("referral_code") or "—"},
    )

    return {
        "invited_count": invited,
        "applied_count": applied,
        "declined_count": declined,
        "apply_rate": apply_rate,
        "decline_rate": decline_rate,
        "avg_match_score": avg_score,
        "top_referral_topics": dict(issues_seen.most_common(8)),
        "referrals_made": referrals_made,
        "referral_code": therapist.get("referral_code"),
    }


@router.get("/portal/therapist/referrals")
async def portal_therapist_referrals(
    request: Request,
    session: dict[str, Any] = Depends(require_session(("therapist",))),
):
    audit.emit(
        actor_type="therapist", actor_id=audit._hash_patient_email(session.get("email", "")),
        action="list_referrals", resource="request", detail="limit=100",
        ip=request.headers.get("x-forwarded-for", ""),
        user_agent=request.headers.get("user-agent", ""),
    )
    therapist = await db.therapists.find_one(
        {"email": {"$regex": f"^{re.escape(session['email'])}$", "$options": "i"}},
        {"_id": 0},
    )
    if not therapist:
        raise HTTPException(404, "Therapist profile not found")

    reqs = await db.requests.find(
        {"notified_therapist_ids": therapist["id"]},
        {"_id": 0, "email": 0, "verification_token": 0},
    ).sort("created_at", -1).to_list(100)

    out = []
    from matching import gap_axes
    for r in reqs:
        score = (r.get("notified_scores") or {}).get(therapist["id"])
        breakdown = (r.get("notified_breakdowns") or {}).get(therapist["id"]) or {}
        gaps = gap_axes(therapist, r, breakdown, top_n=3) if breakdown else []
        application = await db.applications.find_one(
            {"request_id": r["id"], "therapist_id": therapist["id"]},
            {"_id": 0, "id": 1, "message": 1, "created_at": 1,
             "confirms_availability": 1, "confirms_urgency": 1,
             "confirms_payment": 1, "all_confirmed": 1},
        )
        decline = await db.declines.find_one(
            {"request_id": r["id"], "therapist_id": therapist["id"]},
            {"_id": 0},
        )
        if application:
            ref_status = "interested"
        elif decline:
            ref_status = "declined"
        else:
            ref_status = "pending"
        # WS3: server-side tab state
        r_status = (r.get("status") or "").lower()
        if ref_status == "declined":
            state = "past"
        elif ref_status == "interested":
            state = "applied"
        elif r_status in ("delivered", "results_sent", "closed", "archived"):
            state = "past"
        else:
            # 24hr idle expiry -- if the referral has been sitting
            # without a response for >24h, move it to "past".
            matched_ts = r.get("matched_at") or r.get("created_at") or ""
            try:
                matched_dt = datetime.fromisoformat(matched_ts)
                if matched_dt.tzinfo is None:
                    matched_dt = matched_dt.replace(tzinfo=timezone.utc)
                if datetime.now(timezone.utc) - matched_dt > timedelta(hours=24):
                    state = "past"
                else:
                    state = "active"
            except (ValueError, TypeError):
                state = "active"
        out.append({
            "request_id": r["id"],
            "match_score": score,
            "match_breakdown": breakdown,
            "gaps": gaps,
            "created_at": r["created_at"],
            "status": r.get("status"),
            "referral_status": ref_status,
            "state": state,
            "request_status": r_status,
            "deep_match_opt_in": bool(r.get("deep_match_opt_in")),
            "summary": _safe_summary_for_therapist({**r, "email": ""}),
            "presenting_issues_preview": (r.get("presenting_issues") or "")[:140] if isinstance(r.get("presenting_issues"), str) else (", ".join(r.get("presenting_issues") or []))[:140],
            "applied": bool(application),
            "application": application,
            "declined": bool(decline),
        })
    return {
        "therapist": {
            "id": therapist["id"],
            "name": therapist["name"],
            "email": therapist["email"],
            "phone": therapist.get("phone"),
            "phone_alert": therapist.get("phone_alert"),
            "office_phone": therapist.get("office_phone"),
            "credential_type": therapist.get("credential_type"),
            "license_number": therapist.get("license_number"),
            "license_expires_at": therapist.get("license_expires_at"),
            "license_picture": therapist.get("license_picture"),
            "primary_specialties": therapist.get("primary_specialties", []),
            "secondary_specialties": therapist.get("secondary_specialties", []),
            "general_treats": therapist.get("general_treats", []),
            "modalities": therapist.get("modalities", []),
            "modality_offering": therapist.get("modality_offering"),
            "office_locations": therapist.get("office_locations", []),
            "office_addresses": therapist.get("office_addresses", []),
            "website": therapist.get("website"),
            "insurance_accepted": therapist.get("insurance_accepted", []),
            "cash_rate": therapist.get("cash_rate"),
            "sliding_scale": therapist.get("sliding_scale"),
            "free_consult": therapist.get("free_consult"),
            "years_experience": therapist.get("years_experience"),
            "availability_windows": therapist.get("availability_windows", []),
            "urgency_capacity": therapist.get("urgency_capacity"),
            "style_tags": therapist.get("style_tags", []),
            "bio": therapist.get("bio"),
            "profile_picture": therapist.get("profile_picture"),
            "pending_approval": bool(therapist.get("pending_approval")),
            "is_active": therapist.get("is_active", True),
            "availability_prompt_pending": bool(therapist.get("availability_prompt_pending")),
            "last_availability_update_at": therapist.get("last_availability_update_at"),
            "referral_code": therapist.get("referral_code"),
            "pending_reapproval": bool(therapist.get("pending_reapproval")),
            "pending_reapproval_fields": therapist.get("pending_reapproval_fields", []),
            "updated_at": therapist.get("updated_at"),
            "has_password": bool(therapist.get("password_hash")),
            "completeness": _evaluate_profile_inline(therapist),
            # Deep-match T fields. The dashboard chip's
            # _deepMatchMissingLabels() reads these directly off
            # `data.therapist` -- without them in the response shape
            # the chip always counted all 4 as missing even when the
            # therapist had filled them in. (Bug Josh spotted: chip
            # said "4 LEFT" while the edit page showed the answers
            # populated.)
            "t4_hard_truth": therapist.get("t4_hard_truth"),
            "t5_lived_experience": therapist.get("t5_lived_experience"),
            "t6_session_expectations": therapist.get("t6_session_expectations") or [],
            "t6_early_sessions_description": therapist.get("t6_early_sessions_description"),
            # Subscription state surfaced here too so the dashboard
            # ProfileStatusChip + chip CTAs can render the right
            # state without a separate fetch round-trip on every
            # render.
            "subscription_status": therapist.get("subscription_status"),
            "trial_ends_at": therapist.get("trial_ends_at"),
            # Canonical "do they have a real Stripe trial?" indicator.
            # Backfill sets subscription_status="trialing" but leaves
            # stripe_subscription_id as None, so the bare status field
            # can't be trusted. The dashboard chip + edit-page banner
            # both check this presence to decide if a trial is active.
            "stripe_subscription_id": therapist.get("stripe_subscription_id"),
        },
        "referrals": out,
    }


@router.post("/portal/therapist/availability-confirm")
async def portal_therapist_availability_confirm(
    payload: Optional[dict] = None,
    session: dict[str, Any] = Depends(require_session(("therapist",))),
):
    therapist = await db.therapists.find_one(
        {"email": {"$regex": f"^{re.escape(session['email'])}$", "$options": "i"}},
        {"_id": 0, "id": 1},
    )
    if not therapist:
        raise HTTPException(404, "Therapist profile not found")
    update: dict[str, Any] = {
        "availability_prompt_pending": False,
        "last_availability_update_at": _now_iso(),
        "updated_at": _now_iso(),
    }
    if payload and isinstance(payload.get("availability_windows"), list):
        update["availability_windows"] = [
            str(x) for x in payload["availability_windows"] if x
        ][:10]
    if payload and payload.get("urgency_capacity"):
        update["urgency_capacity"] = str(payload["urgency_capacity"])[:50]
    await db.therapists.update_one({"id": therapist["id"]}, {"$set": update})
    return {"ok": True, "updated": list(update.keys())}


# ─── 2FA (TOTP) -- therapist-only ─────────────────────────────────────────────
# Optional second factor for therapist sign-ins. Patients are NOT in scope
# at this stage (their account scope is "see my own intake" -- not worth
# the friction).
#
# Enrollment is 3 steps:
#   1. POST /portal/therapist/2fa/enroll/start  -- get otpauth URI + secret
#   2. POST /portal/therapist/2fa/enroll/verify -- prove the app is set up
#                                                  by entering a current code;
#                                                  receives 10 plaintext
#                                                  recovery codes ONCE
#   3. (frontend stores codes, optionally lets user download/print)
#
# Sign-in change: when therapist has totp_enabled_at set, the existing
# password/magic-code endpoints return {requires_2fa, challenge_token}
# instead of a session token. Client exchanges those at /auth/verify-2fa.

async def _find_therapist_doc(email: str) -> Optional[dict[str, Any]]:
    email_norm = (email or "").lower()
    return await db.therapists.find_one(
        {"email": {"$regex": f"^{re.escape(email_norm)}$", "$options": "i"}},
        {"_id": 0},
    )


@router.post("/portal/therapist/2fa/enroll/start")
async def therapist_2fa_enroll_start(
    session: dict[str, Any] = Depends(require_session(("therapist",))),
):
    """Step 1 of enrollment. Generates a fresh TOTP secret and returns
    the otpauth URI for the frontend to render as a QR code. Does NOT
    persist anything yet -- enrollment is only confirmed once the user
    proves they set up the authenticator correctly via /enroll/verify.

    Returning the same secret on repeated calls would be cleaner UX but
    means a stale tab can hijack enrollment. Forcing a fresh secret on
    each /start call is the safer trade-off.
    """
    from totp_service import generate_secret, otpauth_uri
    therapist = await _find_therapist_doc(session["email"])
    if not therapist:
        raise HTTPException(404, "Therapist profile not found")
    if therapist.get("totp_enabled_at"):
        raise HTTPException(409, "Two-factor authentication is already enabled. Disable it first to re-enroll.")
    secret = generate_secret()
    uri = otpauth_uri(secret, therapist["email"])
    return {"secret": secret, "otpauth_uri": uri}


@router.post("/portal/therapist/2fa/enroll/verify")
async def therapist_2fa_enroll_verify(
    payload: dict[str, Any],
    session: dict[str, Any] = Depends(require_session(("therapist",))),
):
    """Step 2 of enrollment. User submits the secret they were shown +
    the current 6-digit code their authenticator app is displaying. If
    the code verifies, we persist the secret + 10 bcrypt-hashed recovery
    codes, return the recovery codes in plaintext ONCE.

    Body: {"secret": "BASE32...", "code": "123456"}
    """
    from totp_service import (
        verify_code, generate_recovery_codes, hash_recovery_code,
    )
    secret = (payload.get("secret") or "").strip()
    code = (payload.get("code") or "").strip()
    if not secret or not code:
        raise HTTPException(400, "secret and code are required")
    if not verify_code(secret, code):
        raise HTTPException(401, "That code didn't match. Double-check the time on your phone and try the next one (codes refresh every 30 seconds).")
    therapist = await _find_therapist_doc(session["email"])
    if not therapist:
        raise HTTPException(404, "Therapist profile not found")
    if therapist.get("totp_enabled_at"):
        raise HTTPException(409, "Two-factor authentication is already enabled.")
    recovery_plain = generate_recovery_codes()
    recovery_hashes = [hash_recovery_code(c) for c in recovery_plain]
    await db.therapists.update_one(
        {"id": therapist["id"]},
        {"$set": {
            "totp_secret": secret,
            "totp_enabled_at": _now_iso(),
            "totp_recovery_hashes": recovery_hashes,
            "updated_at": _now_iso(),
        }},
    )
    audit.emit(
        actor_type="therapist", actor_id=therapist["id"],
        action="enable_2fa", resource="therapist", resource_id=therapist["id"],
    )
    return {
        "enabled": True,
        "recovery_codes": recovery_plain,
        "enabled_at": _now_iso(),
    }


@router.post("/portal/therapist/2fa/regenerate-recovery")
async def therapist_2fa_regenerate_recovery(
    payload: dict[str, Any],
    session: dict[str, Any] = Depends(require_session(("therapist",))),
):
    """Rotate recovery codes. Requires a fresh valid TOTP code (so a
    stolen session can't silently regenerate them and lock the user out).

    Body: {"code": "123456"}
    """
    from totp_service import (
        verify_code, generate_recovery_codes, hash_recovery_code,
    )
    code = (payload.get("code") or "").strip()
    therapist = await _find_therapist_doc(session["email"])
    if not therapist or not therapist.get("totp_enabled_at"):
        raise HTTPException(404, "2FA is not enabled on this account.")
    if not verify_code(therapist.get("totp_secret", ""), code):
        raise HTTPException(401, "Invalid 2FA code.")
    recovery_plain = generate_recovery_codes()
    recovery_hashes = [hash_recovery_code(c) for c in recovery_plain]
    await db.therapists.update_one(
        {"id": therapist["id"]},
        {"$set": {
            "totp_recovery_hashes": recovery_hashes,
            "updated_at": _now_iso(),
        }},
    )
    audit.emit(
        actor_type="therapist", actor_id=therapist["id"],
        action="regenerate_2fa_recovery", resource="therapist",
        resource_id=therapist["id"],
    )
    return {"recovery_codes": recovery_plain}


@router.post("/portal/therapist/2fa/disable")
async def therapist_2fa_disable(
    payload: dict[str, Any],
    session: dict[str, Any] = Depends(require_session(("therapist",))),
):
    """Turn 2FA off. Requires a fresh valid TOTP code so a stolen
    session can't silently disable it.

    Body: {"code": "123456"}
    """
    from totp_service import verify_code
    code = (payload.get("code") or "").strip()
    therapist = await _find_therapist_doc(session["email"])
    if not therapist or not therapist.get("totp_enabled_at"):
        raise HTTPException(404, "2FA is not enabled on this account.")
    if not verify_code(therapist.get("totp_secret", ""), code):
        raise HTTPException(401, "Invalid 2FA code.")
    await db.therapists.update_one(
        {"id": therapist["id"]},
        {"$unset": {"totp_secret": "", "totp_enabled_at": "",
                    "totp_recovery_hashes": ""},
         "$set": {"updated_at": _now_iso()}},
    )
    audit.emit(
        actor_type="therapist", actor_id=therapist["id"],
        action="disable_2fa", resource="therapist",
        resource_id=therapist["id"],
    )
    return {"enabled": False}


@router.get("/portal/therapist/2fa/status")
async def therapist_2fa_status(
    session: dict[str, Any] = Depends(require_session(("therapist",))),
):
    """Light status check the settings UI uses to render the right
    state (off vs on). Returns enabled_at + recovery_codes_remaining
    when enabled."""
    therapist = await _find_therapist_doc(session["email"])
    if not therapist:
        raise HTTPException(404, "Therapist profile not found")
    enabled = bool(therapist.get("totp_enabled_at"))
    return {
        "enabled": enabled,
        "enabled_at": therapist.get("totp_enabled_at"),
        "recovery_codes_remaining": len(therapist.get("totp_recovery_hashes") or []) if enabled else 0,
    }


@router.post("/auth/verify-2fa")
async def auth_verify_2fa(payload: dict[str, Any], request: Request):
    """Complete the sign-in after the 2FA challenge.

    Body: {
      "challenge_token": "<from /auth/login-password or /auth/verify-code>",
      "code": "<6-digit TOTP code>",
      "use_recovery_code": false,   // when true, `code` is treated as
                                    // a recovery code (XXXX-XXXX-XXXX)
    }
    """
    from totp_service import (
        verify_challenge_token, verify_code, verify_and_consume_recovery_code,
    )
    from login_alerts import check_and_record_login

    challenge = (payload.get("challenge_token") or "").strip()
    code = (payload.get("code") or "").strip()
    use_recovery = bool(payload.get("use_recovery_code"))
    if not challenge or not code:
        raise HTTPException(400, "challenge_token and code are required")
    verified = verify_challenge_token(challenge)
    if not verified:
        raise HTTPException(401, "Sign-in challenge expired. Please sign in again.")
    if verified.get("role") != "therapist":
        # Only therapists have 2FA today; defensive guard.
        raise HTTPException(400, "2FA not configured for this role.")
    email = verified["email"]
    therapist = await _find_therapist_doc(email)
    if not therapist or not therapist.get("totp_enabled_at"):
        # Race: user disabled 2FA between sign-in and verify. Reject
        # so they retry the normal sign-in path.
        raise HTTPException(409, "2FA state changed. Please sign in again.")

    if use_recovery:
        ok, remaining = verify_and_consume_recovery_code(
            code, therapist.get("totp_recovery_hashes") or [],
        )
        if not ok:
            raise HTTPException(401, "That recovery code didn't match.")
        await db.therapists.update_one(
            {"id": therapist["id"]},
            {"$set": {
                "totp_recovery_hashes": remaining,
                "updated_at": _now_iso(),
            }},
        )
    else:
        if not verify_code(therapist.get("totp_secret", ""), code):
            raise HTTPException(401, "That code didn't match. Codes refresh every 30 seconds -- try the next one.")

    token = _create_session_token(email, "therapist")
    await check_and_record_login(
        email=email, role="therapist",
        ip=_client_ip(request),
        user_agent=request.headers.get("user-agent", ""),
    )
    return {
        "token": token,
        "role": "therapist",
        "email": email,
        "has_password": bool(therapist.get("password_hash")),
        "recovery_used": use_recovery,
    }



# ─── Self-serve account deletion ──────────────────────────────────────
# Patient + therapist endpoints that wipe the account on confirmation.
# We require the user to type their email + the literal phrase "DELETE"
# in the request payload so a stolen session can't trigger this in one
# silent click. Audit-logged with the original email so we can prove who
# initiated the request later.
#
# Therapists also get their Stripe subscription cancelled at end-of-period
# (no surprise mid-cycle bills) and their license_document blob cleared.
# Patient deletion wipes their search requests + magic-code history.
#
# We mark documents `deleted_at` + `deleted_reason="self_serve"` rather
# than $unset everything immediately, so the audit trail + 24h
# reversibility window survives. A nightly cron purges anything with
# `deleted_at < now - 24h`.

@router.post("/portal/therapist/delete-account")
async def therapist_delete_account(
    payload: dict[str, Any],
    request: Request,
    session: dict[str, Any] = Depends(require_session(("therapist",))),
):
    """Self-serve account deletion. Requires:
      - confirm_email == session email (case-insensitive)
      - confirm_phrase == "DELETE" (literal, case-sensitive)
    Returns 200 with {"deleted": True} on success."""
    confirm_email = (payload.get("confirm_email") or "").strip().lower()
    confirm_phrase = (payload.get("confirm_phrase") or "").strip()
    if confirm_email != session["email"].lower():
        raise HTTPException(400, "Email confirmation didn't match.")
    if confirm_phrase != "DELETE":
        raise HTTPException(400, 'Type DELETE (all caps) to confirm.')
    therapist = await db.therapists.find_one(
        {"email": {"$regex": f"^{re.escape(session['email'])}$", "$options": "i"}},
        {"_id": 0},
    )
    if not therapist:
        raise HTTPException(404, "Therapist not found")

    # Cancel any active Stripe subscription at period-end so we don't
    # bill again. Best-effort -- if Stripe is down or the sub is already
    # cancelled, log and continue.
    try:
        if therapist.get("stripe_subscription_id"):
            from stripe_service import cancel_subscription
            cancel_subscription(therapist["stripe_subscription_id"])
    except Exception as e:  # noqa: BLE001
        # Cancellation failure shouldn't block deletion; log + move on.
        # Josh can run the manual cancel from admin if needed.
        audit.emit(
            actor_type="therapist", actor_id=therapist["id"],
            action="self_delete_stripe_cancel_failed",
            resource="therapist", resource_id=therapist["id"],
            detail=str(e)[:200],
        )

    now = _now_iso()
    await db.therapists.update_one(
        {"id": therapist["id"]},
        {"$set": {
            "is_active": False,
            "deleted_at": now,
            "deleted_reason": "self_serve",
            "deleted_by_email": session["email"],
            "updated_at": now,
        },
         # Wipe sensitive blobs immediately. Audit + match metadata stay
         # for 24h via the deleted_at flag so a misclick can be reversed
         # by emailing support within the window.
         "$unset": {
             "license_document": "",
             "license_picture": "",
             "totp_secret": "",
             "totp_recovery_hashes": "",
             "password_hash": "",
             "stripe_payment_method_id": "",
         }},
    )
    audit.emit(
        actor_type="therapist", actor_id=therapist["id"],
        action="self_delete_account", resource="therapist",
        resource_id=therapist["id"],
        ip=request.headers.get("x-forwarded-for", ""),
        user_agent=request.headers.get("user-agent", ""),
    )
    return {"deleted": True}


# ─── Self-serve pause / resume ───────────────────────────────────────────────
#
# Pause = stop new patient matches + cancel Stripe sub at period end
# (therapist) or stop matching active requests (patient). Profile,
# license, request history, and feedback-driven reliability scores
# are all preserved -- so algo learning isn't lost. Fully reversible
# any time via /resume; no 24h window like deletion.
#
# Matching filters in helpers.py:_trigger_matching (request-level
# check) and helpers.py:send_matches (therapist-level filter on
# `paused_at: None`) enforce the pause. See admin lifecycle docs for
# the full description of impact on matching + algo learning.

@router.post("/portal/therapist/pause")
async def therapist_pause(
    request: Request,
    session: dict[str, Any] = Depends(require_session(("therapist",))),
):
    """Self-serve pause. Sets paused_at on the therapist doc + cancels
    the Stripe subscription at period end (so no further billing).
    Therapist retains portal access through the existing period and
    can resume at any time."""
    therapist = await db.therapists.find_one(
        {"email": {"$regex": f"^{re.escape(session['email'])}$", "$options": "i"}},
        {"_id": 0},
    )
    if not therapist:
        raise HTTPException(404, "Therapist not found")
    if therapist.get("deleted_at"):
        raise HTTPException(400, "Account is deleted -- can't pause.")
    if therapist.get("paused_at"):
        # Idempotent return; UI may have stale state.
        return {"paused": True, "paused_at": therapist["paused_at"]}

    # Best-effort Stripe cancel at period end. Failure shouldn't block
    # the pause flag itself -- we'd rather have the therapist out of
    # matching than have the pause silently no-op.
    try:
        if therapist.get("stripe_subscription_id"):
            from stripe_service import cancel_subscription
            cancel_subscription(therapist["stripe_subscription_id"])
    except Exception as e:  # noqa: BLE001
        audit.emit(
            actor_type="therapist", actor_id=therapist["id"],
            action="self_pause_stripe_cancel_failed",
            resource="therapist", resource_id=therapist["id"],
            detail=str(e)[:200],
        )

    now = _now_iso()
    await db.therapists.update_one(
        {"id": therapist["id"]},
        {"$set": {
            "paused_at": now,
            "paused_reason": "self_serve",
            "updated_at": now,
        }},
    )
    audit.emit(
        actor_type="therapist", actor_id=therapist["id"],
        action="self_pause_account", resource="therapist",
        resource_id=therapist["id"],
        ip=request.headers.get("x-forwarded-for", ""),
        user_agent=request.headers.get("user-agent", ""),
    )
    return {"paused": True, "paused_at": now}


@router.post("/portal/therapist/resume")
async def therapist_resume(
    request: Request,
    session: dict[str, Any] = Depends(require_session(("therapist",))),
):
    """Resume a paused therapist account. Tries to un-cancel the
    Stripe subscription (if it's still in cancel_at_period_end
    state). If the sub has already fully canceled, returns
    requires_checkout=True so the UI can route the therapist back
    through Stripe checkout to start a new subscription."""
    therapist = await db.therapists.find_one(
        {"email": {"$regex": f"^{re.escape(session['email'])}$", "$options": "i"}},
        {"_id": 0},
    )
    if not therapist:
        raise HTTPException(404, "Therapist not found")
    if therapist.get("deleted_at"):
        raise HTTPException(400, "Account is deleted -- can't resume.")
    if not therapist.get("paused_at"):
        # Already active -- idempotent.
        return {"resumed": True, "requires_checkout": False}

    requires_checkout = False
    try:
        if therapist.get("stripe_subscription_id"):
            from stripe_service import uncancel_subscription
            result = uncancel_subscription(therapist["stripe_subscription_id"])
            if result is None:
                # Sub fully canceled -- need fresh checkout.
                requires_checkout = True
    except Exception as e:  # noqa: BLE001
        audit.emit(
            actor_type="therapist", actor_id=therapist["id"],
            action="self_resume_stripe_uncancel_failed",
            resource="therapist", resource_id=therapist["id"],
            detail=str(e)[:200],
        )
        requires_checkout = True

    now = _now_iso()
    await db.therapists.update_one(
        {"id": therapist["id"]},
        {"$set": {"updated_at": now},
         "$unset": {"paused_at": "", "paused_reason": ""}},
    )
    audit.emit(
        actor_type="therapist", actor_id=therapist["id"],
        action="self_resume_account", resource="therapist",
        resource_id=therapist["id"],
        ip=request.headers.get("x-forwarded-for", ""),
        user_agent=request.headers.get("user-agent", ""),
    )
    return {"resumed": True, "requires_checkout": requires_checkout}


@router.post("/portal/patient/pause")
async def patient_pause(
    request: Request,
    session: dict[str, Any] = Depends(require_session(("patient",))),
):
    """Self-serve patient pause. Marks the patient_accounts row +
    every active request as paused so new matches stop. Existing
    referrals already sent to therapists are not retracted."""
    email_norm = session["email"].lower()
    account = await db.patient_accounts.find_one(
        {"email": email_norm}, {"_id": 0},
    )
    if account and account.get("deleted_at"):
        raise HTTPException(400, "Account is deleted -- can't pause.")
    if account and account.get("paused_at"):
        return {"paused": True, "paused_at": account["paused_at"]}

    now = _now_iso()
    await db.patient_accounts.update_one(
        {"email": email_norm},
        {"$set": {
            "paused_at": now,
            "paused_reason": "self_serve",
            "updated_at": now,
        }},
        upsert=False,
    )
    # Mark every active (non-deleted, non-paused) request paused too,
    # so the matching pipeline skips them. Resume restores them.
    await db.requests.update_many(
        {
            "email": {"$regex": f"^{re.escape(email_norm)}$", "$options": "i"},
            "deleted_at": None,
            "paused_at": None,
        },
        {"$set": {
            "paused_at": now,
            "paused_reason": "patient_self_pause",
            "updated_at": now,
        }},
    )
    audit.emit(
        actor_type="patient",
        actor_id=audit._hash_patient_email(email_norm),
        action="self_pause_account", resource="patient",
        ip=request.headers.get("x-forwarded-for", ""),
        user_agent=request.headers.get("user-agent", ""),
    )
    return {"paused": True, "paused_at": now}


@router.post("/portal/patient/resume")
async def patient_resume(
    request: Request,
    session: dict[str, Any] = Depends(require_session(("patient",))),
):
    """Resume a paused patient account. Un-pauses the account row +
    every request that was paused by the same self-serve action
    (paused_reason='patient_self_pause'). Requests paused by other
    actors -- e.g. admin freeze -- are left alone."""
    email_norm = session["email"].lower()
    account = await db.patient_accounts.find_one(
        {"email": email_norm}, {"_id": 0},
    )
    if account and account.get("deleted_at"):
        raise HTTPException(400, "Account is deleted -- can't resume.")
    if not account or not account.get("paused_at"):
        return {"resumed": True}

    now = _now_iso()
    await db.patient_accounts.update_one(
        {"email": email_norm},
        {"$set": {"updated_at": now},
         "$unset": {"paused_at": "", "paused_reason": ""}},
    )
    await db.requests.update_many(
        {
            "email": {"$regex": f"^{re.escape(email_norm)}$", "$options": "i"},
            "paused_reason": "patient_self_pause",
        },
        {"$set": {"updated_at": now},
         "$unset": {"paused_at": "", "paused_reason": ""}},
    )
    audit.emit(
        actor_type="patient",
        actor_id=audit._hash_patient_email(email_norm),
        action="self_resume_account", resource="patient",
        ip=request.headers.get("x-forwarded-for", ""),
        user_agent=request.headers.get("user-agent", ""),
    )
    return {"resumed": True}


# ─── Self-serve data export ──────────────────────────────────────────────────
#
# Both roles can download a JSON snapshot of their account data:
# profile + activity history + audit-relevant breadcrumbs. Sensitive
# auth artifacts (password_hash, totp_secret, totp_recovery_hashes,
# magic codes) are stripped before the export hits the wire.
#
# Returns application/json with Content-Disposition: attachment so
# the browser triggers a download. The download link itself is
# session-authenticated -- only the account owner can pull their
# own data.

_THERAPIST_EXPORT_STRIP = {
    "_id", "password_hash", "password_set_at", "totp_secret",
    "totp_recovery_hashes", "t5_embedding", "t6b_embedding",
    "license_document",  # base64 blob -- huge, and patient never
                         # uploaded it; the therapist already has
                         # the original.
}
_PATIENT_ACCOUNT_STRIP = {
    "_id", "password_hash", "password_set_at", "totp_secret",
    "totp_recovery_hashes",
}
_REQUEST_STRIP = {"_id", "verification_token"}
_APPLICATION_STRIP = {"_id"}
_FEEDBACK_STRIP = {"_id"}
_LOGIN_EVENT_STRIP = {"_id"}


def _strip_keys(doc: dict[str, Any], keys: set[str]) -> dict[str, Any]:
    return {k: v for k, v in (doc or {}).items() if k not in keys}


def _export_filename(role: str, email: str) -> str:
    safe_email = re.sub(r"[^a-zA-Z0-9._-]+", "_", email)[:64] or "account"
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    return f"theravoca-{role}-{safe_email}-{today}.xlsx"


def _flatten_for_excel(value: Any) -> str:
    """Convert nested dicts/lists/dates into a single cell-friendly
    string. Excel can't represent nested structures natively, so we
    serialize them as compact JSON for the user to read."""
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return value if not isinstance(value, bool) else ("Yes" if value else "No")
    if isinstance(value, (list, dict)):
        try:
            return json.dumps(value, default=str, ensure_ascii=False)
        except (TypeError, ValueError):
            return str(value)
    return str(value)


def _build_excel_workbook(sheets: list[tuple[str, list[dict]]]) -> bytes:
    """Build a multi-sheet xlsx workbook from a list of (sheet_name, rows)
    tuples. Each rows list is a list of dicts; the union of keys becomes
    the column headers. Returns the raw .xlsx bytes."""
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    wb = Workbook()
    # Remove the auto-created default sheet so we control all sheet names.
    if wb.worksheets:
        wb.remove(wb.active)

    for sheet_name, rows in sheets:
        # Excel sheet names cap at 31 chars and can't contain []:*?/\
        clean_name = re.sub(r"[\[\]:*?/\\]", "_", sheet_name)[:31] or "Sheet"
        ws = wb.create_sheet(clean_name)
        if not rows:
            ws.append(["(no records)"])
            continue
        # Column headers = union of all keys across rows, sorted for
        # stable output.
        all_keys = sorted({k for row in rows for k in (row or {}).keys()})
        ws.append(all_keys)
        for row in rows:
            ws.append([_flatten_for_excel((row or {}).get(k)) for k in all_keys])
        # Best-effort column widths: cap at 60 chars so long JSON blobs
        # don't make the sheet unreadable. Excel auto-wraps within the
        # cell when the user widens it manually.
        for idx, key in enumerate(all_keys, start=1):
            max_len = max(
                [len(str(key))] + [len(str((r or {}).get(key) or "")) for r in rows[:50]]
            )
            ws.column_dimensions[get_column_letter(idx)].width = min(max(12, max_len + 2), 60)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


_XLSX_MEDIA_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)


@router.get("/portal/therapist/export-data")
async def therapist_export_data(
    request: Request,
    session: dict[str, Any] = Depends(require_session(("therapist",))),
):
    """Download a JSON snapshot of the therapist's account: profile,
    referrals (applications + declines), feedback ABOUT them, and
    sign-in history. Auth artifacts and internal scoring caches are
    stripped."""
    email = session["email"]
    therapist = await db.therapists.find_one(
        {"email": {"$regex": f"^{re.escape(email)}$", "$options": "i"}},
    )
    if not therapist:
        raise HTTPException(404, "Therapist not found")
    tid = therapist["id"]

    apps = await db.applications.find(
        {"therapist_id": tid}, {"_id": 0},
    ).to_list(2000)
    declines = await db.declines.find(
        {"therapist_id": tid}, {"_id": 0},
    ).to_list(2000)
    fb = await db.feedback.find(
        {"therapist_id": tid}, {"_id": 0},
    ).to_list(2000)
    logins = await db.login_events.find(
        {"role": "therapist", "email": email.lower()},
        {"_id": 0},
    ).sort("at", -1).to_list(500)

    # Build an Excel workbook with one sheet per data class so the user
    # can scan their data in Numbers/Excel/Sheets without parsing JSON.
    # Each sheet's columns are the union of keys; nested values
    # (lists/dicts/dates) are JSON-serialized into a single cell.
    sheets = [
        ("Summary", [{
            "exported_at": _now_iso(),
            "exported_for": email,
            "role": "therapist",
        }]),
        ("Profile", [_strip_keys(therapist, _THERAPIST_EXPORT_STRIP)]),
        ("Applications", [_strip_keys(a, _APPLICATION_STRIP) for a in apps]),
        ("Declines", declines),
        ("Feedback About Me", [_strip_keys(f, _FEEDBACK_STRIP) for f in fb]),
        ("Login History", [_strip_keys(le, _LOGIN_EVENT_STRIP) for le in logins]),
    ]
    content = _build_excel_workbook(sheets)
    audit.emit(
        actor_type="therapist", actor_id=tid,
        action="self_export_data", resource="therapist",
        resource_id=tid,
        ip=request.headers.get("x-forwarded-for", ""),
        user_agent=request.headers.get("user-agent", ""),
    )
    return Response(
        content=content,
        media_type=_XLSX_MEDIA_TYPE,
        headers={
            "Content-Disposition": (
                f'attachment; filename="{_export_filename("therapist", email)}"'
            ),
        },
    )


@router.get("/portal/patient/export-data")
async def patient_export_data(
    request: Request,
    session: dict[str, Any] = Depends(require_session(("patient",))),
):
    """Download a JSON snapshot of the patient's account: account
    profile, every match request they've submitted, the therapist
    applications/replies on those requests, feedback the patient
    submitted, and sign-in history. Auth artifacts stripped."""
    email = session["email"]
    email_lower = email.lower()
    account = await db.patient_accounts.find_one({"email": email_lower})
    requests_docs = await db.requests.find(
        {"email": {"$regex": f"^{re.escape(email)}$", "$options": "i"}},
        {"_id": 0, "verification_token": 0},
    ).to_list(500)
    request_ids = [r["id"] for r in requests_docs if r.get("id")]
    apps = []
    if request_ids:
        apps = await db.applications.find(
            {"request_id": {"$in": request_ids}}, {"_id": 0},
        ).to_list(2000)
    fb = await db.feedback.find(
        {"patient_email": email_lower}, {"_id": 0},
    ).to_list(2000)
    logins = await db.login_events.find(
        {"role": "patient", "email": email_lower}, {"_id": 0},
    ).sort("at", -1).to_list(500)

    sheets = [
        ("Summary", [{
            "exported_at": _now_iso(),
            "exported_for": email,
            "role": "patient",
        }]),
        ("Account", [_strip_keys(account or {}, _PATIENT_ACCOUNT_STRIP)]),
        ("Match Requests", [_strip_keys(r, _REQUEST_STRIP) for r in requests_docs]),
        ("Therapist Replies", [_strip_keys(a, _APPLICATION_STRIP) for a in apps]),
        ("Feedback I Submitted", [_strip_keys(f, _FEEDBACK_STRIP) for f in fb]),
        ("Login History", [_strip_keys(le, _LOGIN_EVENT_STRIP) for le in logins]),
    ]
    content = _build_excel_workbook(sheets)
    audit.emit(
        actor_type="patient",
        actor_id=audit._hash_patient_email(email_lower),
        action="self_export_data", resource="patient",
        ip=request.headers.get("x-forwarded-for", ""),
        user_agent=request.headers.get("user-agent", ""),
    )
    return Response(
        content=content,
        media_type=_XLSX_MEDIA_TYPE,
        headers={
            "Content-Disposition": (
                f'attachment; filename="{_export_filename("patient", email)}"'
            ),
        },
    )


@router.post("/portal/patient/delete-account")
async def patient_delete_account(
    payload: dict[str, Any],
    request: Request,
    session: dict[str, Any] = Depends(require_session(("patient",))),
):
    """Self-serve patient account deletion. Same confirmation gate as
    the therapist endpoint."""
    confirm_email = (payload.get("confirm_email") or "").strip().lower()
    confirm_phrase = (payload.get("confirm_phrase") or "").strip()
    if confirm_email != session["email"].lower():
        raise HTTPException(400, "Email confirmation didn't match.")
    if confirm_phrase != "DELETE":
        raise HTTPException(400, 'Type DELETE (all caps) to confirm.')

    email_norm = session["email"].lower()
    now = _now_iso()
    # Mark patient_accounts row deleted (24h reversibility window)
    await db.patient_accounts.update_one(
        {"email": email_norm},
        {"$set": {
            "deleted_at": now,
            "deleted_reason": "self_serve",
            "updated_at": now,
        },
         "$unset": {"password_hash": "", "totp_secret": "",
                    "totp_recovery_hashes": ""}},
        upsert=False,
    )
    # Mark all the patient's search requests deleted too -- they
    # shouldn't appear in matching after the patient deletes.
    await db.requests.update_many(
        {"email": {"$regex": f"^{re.escape(email_norm)}$", "$options": "i"}},
        {"$set": {
            "deleted_at": now,
            "deleted_reason": "patient_self_delete",
            "is_active": False,
            "updated_at": now,
        }},
    )
    audit.emit(
        actor_type="patient",
        actor_id=audit._hash_patient_email(email_norm),
        action="self_delete_account", resource="patient",
        ip=request.headers.get("x-forwarded-for", ""),
        user_agent=request.headers.get("user-agent", ""),
    )
    return {"deleted": True}
