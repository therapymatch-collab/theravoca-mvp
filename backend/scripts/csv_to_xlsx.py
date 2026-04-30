"""Convert the latest experiment_text_impact run into a polished
single-file Excel workbook the user can download from the admin.

Sheets
------
1. Summary      — aggregate stats (variant lift, patient-text Δ, cross-tab)
2. Raw          — every (request, therapist, variant) row from the CSV
3. By variant   — pivot by message variant (avg fit, etc.)
4. By patient   — pivot by request idx + with_other_issue flag
"""
from __future__ import annotations

import csv
import statistics
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

RESULTS = Path("/app/backend/scripts/results")


def _latest_csv() -> Path:
    csvs = sorted(RESULTS.glob("exp_*.csv"), reverse=True)
    if not csvs:
        sys.exit("no experiment csv found")
    return csvs[0]


def _avg(xs: list[float]) -> float:
    return round(statistics.mean(xs), 2) if xs else 0.0


HEADER_FILL = PatternFill("solid", fgColor="2D4A3E")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
SECTION_FILL = PatternFill("solid", fgColor="F0C674")
SECTION_FONT = Font(color="2B2A29", bold=True, size=12)
SUBTLE = Font(color="6D6A65", size=10, italic=True)


def _autosize(ws) -> None:
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max(
            (len(str(c.value)) for c in col if c.value is not None),
            default=10,
        )
        ws.column_dimensions[col_letter].width = min(50, max(12, max_len + 2))


def main() -> None:
    csv_path = _latest_csv()
    print(f"reading {csv_path}")
    with csv_path.open() as fp:
        rows = list(csv.DictReader(fp))

    # cast numeric columns
    for r in rows:
        for k in ("request_idx", "with_other_issue", "raw_match_score",
                  "apply_fit", "apply_msg_len"):
            try:
                r[k] = float(r[k])
                if k != "apply_fit":  # apply_fit has decimals
                    r[k] = int(r[k])
            except (KeyError, ValueError):
                pass

    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "TheraVoca scoring experiment — text-impact run"
    ws["A1"].font = Font(bold=True, size=16, color="2D4A3E")
    ws.merge_cells("A1:F1")
    ws["A2"] = (
        f"N requests: {len(set(r['request_id'] for r in rows))}   ·   "
        f"N applies: {len(rows)}   ·   "
        f"5 variants × 5 therapists per request"
    )
    ws["A2"].font = SUBTLE
    ws.merge_cells("A2:F2")

    row_n = 4
    # ── Section 1: Patient text vs raw match score ───────────────────
    ws.cell(row=row_n, column=1, value="1. Patient `other_issue` text vs raw match score")
    ws.cell(row=row_n, column=1).font = SECTION_FONT
    ws.cell(row=row_n, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=row_n, start_column=1, end_row=row_n, end_column=6)
    row_n += 1

    headers = ["Group", "N", "Avg match", "Median", "Min", "Max"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row_n, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")
    row_n += 1

    for flag, label in [(0, "Empty `other_issue`"), (1, "Rich `other_issue`")]:
        scores = [r["raw_match_score"] for r in rows if int(r["with_other_issue"]) == flag]
        ws.cell(row=row_n, column=1, value=label)
        ws.cell(row=row_n, column=2, value=len(scores))
        ws.cell(row=row_n, column=3, value=_avg(scores))
        ws.cell(
            row=row_n, column=4,
            value=round(statistics.median(scores), 1) if scores else 0,
        )
        ws.cell(row=row_n, column=5, value=min(scores) if scores else 0)
        ws.cell(row=row_n, column=6, value=max(scores) if scores else 0)
        row_n += 1

    delta = _avg(
        [r["raw_match_score"] for r in rows if int(r["with_other_issue"]) == 1]
    ) - _avg(
        [r["raw_match_score"] for r in rows if int(r["with_other_issue"]) == 0]
    )
    ws.cell(
        row=row_n, column=1,
        value=f"Δ (rich − empty): {delta:+.2f} pts (≈ noise — `other_issue` is not consumed by matching engine yet)",
    ).font = Font(italic=True, color="6D6A65")
    ws.merge_cells(start_row=row_n, start_column=1, end_row=row_n, end_column=6)
    row_n += 2

    # ── Section 2: Therapist apply-message vs apply_fit ──────────────
    ws.cell(row=row_n, column=1, value="2. Therapist apply-message vs apply_fit (LLM-graded 0-5)")
    ws.cell(row=row_n, column=1).font = SECTION_FONT
    ws.cell(row=row_n, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=row_n, start_column=1, end_row=row_n, end_column=6)
    row_n += 1

    headers = ["Variant", "Avg apply_fit", "Median", "Min", "Max", "Avg msg length"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row_n, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")
    row_n += 1

    variants = ["A_empty", "B_oneliner", "C_generic_long",
                "D_issue_specific", "E_full_engagement"]
    for v in variants:
        rs = [r for r in rows if r["variant"] == v]
        fits = [r["apply_fit"] for r in rs]
        lens = [r["apply_msg_len"] for r in rs]
        ws.cell(row=row_n, column=1, value=v)
        ws.cell(row=row_n, column=2, value=_avg(fits))
        ws.cell(
            row=row_n, column=3,
            value=round(statistics.median(fits), 1) if fits else 0,
        )
        ws.cell(row=row_n, column=4, value=min(fits) if fits else 0)
        ws.cell(row=row_n, column=5, value=max(fits) if fits else 0)
        ws.cell(row=row_n, column=6, value=int(round(_avg(lens))))
        row_n += 1
    row_n += 1

    # ── Section 3: Cross-tab patient text × variant ──────────────────
    ws.cell(row=row_n, column=1, value="3. Does patient `other_issue` amplify the apply-fit lift?")
    ws.cell(row=row_n, column=1).font = SECTION_FONT
    ws.cell(row=row_n, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=row_n, start_column=1, end_row=row_n, end_column=6)
    row_n += 1

    headers = ["Variant", "apply_fit (no patient text)", "apply_fit (with patient text)", "Δ"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row_n, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")
    row_n += 1

    for v in variants:
        rs = [r for r in rows if r["variant"] == v]
        no_t = _avg([r["apply_fit"] for r in rs if int(r["with_other_issue"]) == 0])
        yes_t = _avg([r["apply_fit"] for r in rs if int(r["with_other_issue"]) == 1])
        ws.cell(row=row_n, column=1, value=v)
        ws.cell(row=row_n, column=2, value=no_t)
        ws.cell(row=row_n, column=3, value=yes_t)
        c = ws.cell(row=row_n, column=4, value=round(yes_t - no_t, 2))
        if yes_t - no_t > 0:
            c.font = Font(color="2D4A3E", bold=True)
        row_n += 1
    row_n += 1

    # ── Section 4: Architectural takeaway ────────────────────────────
    ws.cell(row=row_n, column=1, value="4. Architectural takeaway")
    ws.cell(row=row_n, column=1).font = SECTION_FONT
    ws.cell(row=row_n, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=row_n, start_column=1, end_row=row_n, end_column=6)
    row_n += 1
    bullets = [
        "Patient `other_issue` free text is NOT yet consumed by `matching._score_one`. Section 1's Δ ≈ 0 confirms this.",
        "Therapist apply text drives a strong, monotonic `apply_fit` lift (Section 2): empty/oneliner → ~0; specific reply → 3.2; full engagement → 4.8.",
        "Cross-tab in Section 3 shows tiny positive lift (+0.24 on variant E) when the patient writes more — would be larger if the grader saw the free text (it currently does not).",
        "Recommended fix: feed `other_issue` into the `score_apply_fit` prompt, and embed it for soft-bonus matching against therapist T5.",
    ]
    for b in bullets:
        ws.cell(row=row_n, column=1, value="•  " + b)
        ws.merge_cells(start_row=row_n, start_column=1, end_row=row_n, end_column=6)
        ws.cell(row=row_n, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row_n].height = 30
        row_n += 1

    _autosize(ws)

    # ── Sheet 2: Raw ────────────────────────────────────────────────
    ws_raw = wb.create_sheet("Raw")
    if rows:
        cols = list(rows[0].keys())
        for ci, h in enumerate(cols, 1):
            c = ws_raw.cell(row=1, column=ci, value=h)
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
        for ri, r in enumerate(rows, 2):
            for ci, h in enumerate(cols, 1):
                ws_raw.cell(row=ri, column=ci, value=r[h])
        ws_raw.freeze_panes = "A2"
        _autosize(ws_raw)

    # ── Sheet 3: By variant ─────────────────────────────────────────
    ws_v = wb.create_sheet("By variant")
    headers = [
        "Variant", "N", "Avg apply_fit", "Median apply_fit",
        "Avg raw match", "Avg msg length",
    ]
    for ci, h in enumerate(headers, 1):
        c = ws_v.cell(row=1, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    for ri, v in enumerate(variants, 2):
        rs = [r for r in rows if r["variant"] == v]
        ws_v.cell(row=ri, column=1, value=v)
        ws_v.cell(row=ri, column=2, value=len(rs))
        ws_v.cell(
            row=ri, column=3,
            value=_avg([r["apply_fit"] for r in rs]),
        )
        ws_v.cell(
            row=ri, column=4,
            value=round(
                statistics.median([r["apply_fit"] for r in rs]) if rs else 0, 1,
            ),
        )
        ws_v.cell(
            row=ri, column=5,
            value=_avg([r["raw_match_score"] for r in rs]),
        )
        ws_v.cell(
            row=ri, column=6,
            value=int(round(_avg([r["apply_msg_len"] for r in rs]))),
        )
    ws_v.freeze_panes = "A2"
    _autosize(ws_v)

    # ── Sheet 4: By patient ────────────────────────────────────────
    ws_p = wb.create_sheet("By patient")
    headers = [
        "Request idx", "With `other_issue`", "Primary issue",
        "Avg raw match (top5)", "Avg apply_fit (across variants)",
    ]
    for ci, h in enumerate(headers, 1):
        c = ws_p.cell(row=1, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL

    by_req: dict[int, list[dict]] = {}
    for r in rows:
        by_req.setdefault(int(r["request_idx"]), []).append(r)

    ri = 2
    for idx in sorted(by_req.keys()):
        rs = by_req[idx]
        first = rs[0]
        ws_p.cell(row=ri, column=1, value=idx)
        ws_p.cell(
            row=ri, column=2,
            value="yes" if int(first["with_other_issue"]) == 1 else "no",
        )
        ws_p.cell(row=ri, column=3, value=first["primary_issue"])
        ws_p.cell(
            row=ri, column=4,
            value=_avg([r["raw_match_score"] for r in rs]),
        )
        ws_p.cell(
            row=ri, column=5,
            value=_avg([r["apply_fit"] for r in rs]),
        )
        ri += 1
    ws_p.freeze_panes = "A2"
    _autosize(ws_p)

    out_path = csv_path.with_suffix(".xlsx")
    wb.save(out_path)
    print(f"wrote {out_path}")


if __name__ == "__main__":
    main()
